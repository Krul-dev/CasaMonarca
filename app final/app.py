# correr con 
#$env:DB_USER="root"; $env:DB_PASSWORD="nueva123"; python app.py

import io
import os
import base64
import re
import secrets
import string
import pymysql
import pymysql.cursors
import glob
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from functools import wraps
from datetime import datetime, timedelta

from cryptography import x509
from cryptography.x509.oid import NameOID
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import rsa, ec
from cryptography.hazmat.primitives.serialization import pkcs12

from werkzeug.security import generate_password_hash, check_password_hash

from flask import (
    Flask, render_template, request, redirect,
    url_for, session, g, abort, send_file, jsonify
)
from flask_session import Session

# ── CONFIGURACIÓN ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'demo-secret-key-cambiar-en-produccion'

SESSION_DIR  = os.path.join(os.getcwd(), 'flask_sessions_demo')
CA_DIR       = os.path.join(os.getcwd(), 'ca')
DB_DIR       = os.path.join(os.getcwd(), 'db')
SCHEMA_PATH  = os.path.join(DB_DIR, 'schema_mysql.sql')
DB_HOST      = os.environ.get('DB_HOST', 'localhost')
DB_PORT      = int(os.environ.get('DB_PORT', '3306'))
DB_USER      = os.environ.get('DB_USER', 'root')
DB_PASSWORD  = os.environ.get('DB_PASSWORD', '')
DB_NAME      = os.environ.get('DB_NAME', 'casa_monarca')
CA_CERT_PATH = os.path.join(CA_DIR, 'ca_cert.pem')
CA_KEY_PATH  = os.path.join(CA_DIR, 'ca_key.pem')
CA_PASSWORD  = os.environ.get('CA_PASSWORD', 'demo-ca-pwd-cambiar').encode('utf-8')

for d in (SESSION_DIR, CA_DIR, DB_DIR):
    os.makedirs(d, exist_ok=True)

app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = SESSION_DIR
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['TEMPLATES_AUTO_RELOAD'] = True
Session(app)

# ── SEEDS INICIALES (se vuelcan a la DB la primera vez) ───────
USUARIOS_DEMO = {
    'admin@casamonarca.org':    {'nombre': 'Administrador Principal', 'rol': 'admin',  'password': 'admin123',    'activo': 1},
    'respaldo@casamonarca.org': {'nombre': 'Admin Respaldo',          'rol': 'admin',  'password': 'Respaldo@25', 'activo': 0},
    'coord@casamonarca.org':    {'nombre': 'Coordinadora Demo',       'rol': 'coord',  'password': 'coord123',    'activo': 1},
}

CERTIFICADOS_SEED = [
    {
        'serial': '4A2F88C1E3',
        'usuario': 'Ana García López',
        'rol': 'op',
        'fecha_emision': '2025-04-01',
        'activo': True,
    },
    {
        'serial': '3E1D77B2F9',
        'usuario': 'Luis Martínez',
        'rol': 'coord',
        'fecha_emision': '2025-03-22',
        'activo': False,
    },
    {
        'serial': '7C4A19D0B5',
        'usuario': 'Sofía Ramírez',
        'rol': 'op',
        'fecha_emision': '2025-03-10',
        'activo': True,
    },
]

# ── AUTORIDAD CERTIFICADORA (persistente en disco) ────────────
# Se genera UNA sola vez la primera vez que arranca el servidor.
# A partir de ahí, todos los certificados de voluntarios quedan
# firmados por esta misma CA (antes eran autofirmados).
def inicializar_ca():
    if os.path.exists(CA_CERT_PATH) and os.path.exists(CA_KEY_PATH):
        return
    ca_key = rsa.generate_private_key(public_exponent=65537, key_size=4096)
    ca_name = x509.Name([
        x509.NameAttribute(NameOID.COUNTRY_NAME, 'MX'),
        x509.NameAttribute(NameOID.ORGANIZATION_NAME, 'Casa Monarca'),
        x509.NameAttribute(NameOID.COMMON_NAME, 'CA Casa Monarca'),
    ])
    ca_cert = (
        x509.CertificateBuilder()
        .subject_name(ca_name)
        .issuer_name(ca_name)
        .public_key(ca_key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(datetime.utcnow() - timedelta(days=1))
        .not_valid_after(datetime.utcnow() + timedelta(days=3650))
        .add_extension(x509.BasicConstraints(ca=True, path_length=None), critical=True)
        .sign(ca_key, hashes.SHA256())
    )
    with open(CA_CERT_PATH, 'wb') as f:
        f.write(ca_cert.public_bytes(serialization.Encoding.PEM))
    with open(CA_KEY_PATH, 'wb') as f:
        f.write(ca_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.BestAvailableEncryption(CA_PASSWORD)
        ))

def cargar_ca():
    with open(CA_CERT_PATH, 'rb') as f:
        ca_cert = x509.load_pem_x509_certificate(f.read())
    with open(CA_KEY_PATH, 'rb') as f:
        ca_key = serialization.load_pem_private_key(f.read(), password=CA_PASSWORD)
    return ca_cert, ca_key

# ── CERTIFICADO DEL SERVIDOR (para TLS / mTLS) ────────────────
# Genera un cert de servidor firmado por la CA la primera vez.
# Este es el cert que el navegador ve del lado del servidor.
SERVER_CERT_PATH = os.path.join(CA_DIR, 'server.pem')
SERVER_KEY_PATH  = os.path.join(CA_DIR, 'server.key')

def inicializar_cert_servidor():
    if os.path.exists(SERVER_CERT_PATH) and os.path.exists(SERVER_KEY_PATH):
        return
    ca_cert, ca_key = cargar_ca()
    srv_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    subject = x509.Name([
        x509.NameAttribute(NameOID.COUNTRY_NAME, 'MX'),
        x509.NameAttribute(NameOID.ORGANIZATION_NAME, 'Casa Monarca'),
        x509.NameAttribute(NameOID.COMMON_NAME, 'localhost'),
    ])
    srv_cert = (
        x509.CertificateBuilder()
        .subject_name(subject)
        .issuer_name(ca_cert.subject)
        .public_key(srv_key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(datetime.utcnow() - timedelta(days=1))
        .not_valid_after(datetime.utcnow() + timedelta(days=3650))
        .add_extension(x509.SubjectAlternativeName([
            x509.DNSName('localhost'),
            x509.DNSName('127.0.0.1'),
        ]), critical=False)
        .add_extension(x509.BasicConstraints(ca=False, path_length=None), critical=True)
        .sign(ca_key, hashes.SHA256())
    )
    with open(SERVER_CERT_PATH, 'wb') as f:
        f.write(srv_cert.public_bytes(serialization.Encoding.PEM))
    with open(SERVER_KEY_PATH, 'wb') as f:
        f.write(srv_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.NoEncryption()
        ))

# ── BASE DE DATOS (MySQL) ─────────────────────────────────────
# ── FIRMAS ECDSA (P-256) ──────────────────────────────────────
def generar_claves_ec():
    privada = ec.generate_private_key(ec.SECP256R1())
    priv_pem = privada.private_bytes(
        serialization.Encoding.PEM,
        serialization.PrivateFormat.PKCS8,
        serialization.NoEncryption()
    ).decode()
    pub_pem = privada.public_key().public_bytes(
        serialization.Encoding.PEM,
        serialization.PublicFormat.SubjectPublicKeyInfo
    ).decode()
    return priv_pem, pub_pem

def _obtener_o_crear_claves_ec(db, usuario):
    row = _exec(db, 'SELECT ec_private_key, ec_public_key FROM usuarios WHERE usuario=%s',
                (usuario,)).fetchone()
    if row and row['ec_private_key']:
        return row['ec_private_key'], row['ec_public_key']
    priv_pem, pub_pem = generar_claves_ec()
    _exec(db, 'UPDATE usuarios SET ec_private_key=%s, ec_public_key=%s WHERE usuario=%s',
          (priv_pem, pub_pem, usuario))
    db.commit()
    return priv_pem, pub_pem

def firmar_mensaje(priv_pem: str, mensaje: str) -> str:
    privada = serialization.load_pem_private_key(priv_pem.encode(), password=None)
    sig = privada.sign(mensaje.encode('utf-8'), ec.ECDSA(hashes.SHA256()))
    return base64.b64encode(sig).decode()

def verificar_firma(pub_pem: str, mensaje: str, firma_b64: str) -> bool:
    try:
        publica = serialization.load_pem_public_key(pub_pem.encode())
        publica.verify(base64.b64decode(firma_b64),
                       mensaje.encode('utf-8'), ec.ECDSA(hashes.SHA256()))
        return True
    except Exception:
        return False

def obtener_db():
    db = getattr(g, '_db', None)
    if db is None:
        g._db = pymysql.connect(
            host=DB_HOST, port=DB_PORT,
            user=DB_USER, password=DB_PASSWORD,
            database=DB_NAME,
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=False
        )
    return g._db

@app.teardown_appcontext
def cerrar_db(_exc):
    db = getattr(g, '_db', None)
    if db is not None:
        db.close()

def _exec(db, sql, params=None):
    cur = db.cursor()
    cur.execute(sql, params or ())
    return cur

def inicializar_db():
    """Crea tablas y siembra USUARIOS_DEMO + CERTIFICADOS_SEED la
    primera vez que arranca la app."""
    con = pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASSWORD,
        database=DB_NAME,
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False
    )
    with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
        schema = f.read()
    cur = con.cursor()
    for stmt in schema.split(';'):
        stmt = stmt.strip()
        if stmt:
            cur.execute(stmt)
    cur.close()

    # Migración: garantizar que usuarios.rol incluya 'voluntario'.
    try:
        cur = con.cursor()
        cur.execute(
            "ALTER TABLE usuarios MODIFY rol "
            "ENUM('admin','coord','op','voluntario') NOT NULL"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración usuarios.rol ENUM] {e}')

    # Migración: log_certificados.resultado — incluir todos los valores actuales.
    try:
        cur = con.cursor()
        cur.execute(
            "ALTER TABLE log_certificados MODIFY resultado "
            "ENUM('login_exitoso','login_fallido','pwd_incorrecta',"
            "     'cert_revocado','cert_expirado','cert_emitido',"
            "     'cert_revocado_admin','rol_modificado','pwd_cambiada',"
            "     'acceso_denegado','migrante_registrado','migrante_actualizado',"
            "     'migrante_solicitud_eliminacion','migrante_eliminado',"
            "     'pwd_reset_solicitado','pwd_reset_aprobado','pwd_reset_rechazado') NOT NULL"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración log resultado ENUM] {e}')

    # Migración: agregar campo activo a usuarios si no existe.
    try:
        cur = con.cursor()
        cur.execute(
            "ALTER TABLE usuarios ADD COLUMN activo TINYINT(1) NOT NULL DEFAULT 1"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración usuarios.activo] {e}')

    # Migración: crear tabla solicitudes_pwd si no existe (ya está en schema, esto es por seguridad).
    try:
        cur = con.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS solicitudes_pwd ("
            "  id            INT AUTO_INCREMENT PRIMARY KEY,"
            "  usuario       VARCHAR(120) NOT NULL,"
            "  estado        ENUM('pendiente','aprobada','rechazada') DEFAULT 'pendiente',"
            "  solicitado_en DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,"
            "  resuelto_por  VARCHAR(64),"
            "  resuelto_en   DATETIME,"
            "  INDEX idx_spwd_usuario (usuario),"
            "  INDEX idx_spwd_estado  (estado)"
            ") ENGINE=InnoDB"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración solicitudes_pwd] {e}')

    # Migración: nuevos campos del cuestionario de migrantes.
    nuevas_cols_migrantes = [
        ("fecha_atencion",     "DATE"),
        ("genero",             "VARCHAR(30)"),
        ("departamento_estado","VARCHAR(100)"),
        ("estado_civil",       "VARCHAR(30)"),
        ("grupo_poblacion",    "VARCHAR(80)"),
        ("telefono_contacto",  "VARCHAR(32)"),
    ]
    for col, tipo in nuevas_cols_migrantes:
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE migrantes ADD COLUMN {col} {tipo}")
            con.commit()
            cur.close()
        except Exception as e:
            print(f'[migración migrantes.{col}] {e}')

    # Migración: eliminar columnas obsoletas de migrantes (reemplazadas por el nuevo formulario).
    cols_obsoletas_migrantes = [
        'tipo_documento', 'num_documento', 'nacionalidad', 'estado_migratorio',
        'fecha_ingreso', 'fecha_egreso', 'tiene_pasaporte', 'tiene_visa',
        'tiene_identificacion', 'estado_salud', 'obs_medicas',
        'contacto_emergencia', 'telefono_emergencia', 'observaciones',
    ]
    for col in cols_obsoletas_migrantes:
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE migrantes DROP COLUMN {col}")
            con.commit()
            cur.close()
        except Exception as e:
            pass  # columna ya eliminada o no existe

    # Migración: corregir tipos de columnas nuevas de migrantes (por si quedaron mal definidas).
    cols_tipos_migrantes = [
        ('genero',             'VARCHAR(30)'),
        ('departamento_estado','VARCHAR(100)'),
        ('estado_civil',       'VARCHAR(30)'),
        ('grupo_poblacion',    'VARCHAR(80)'),
        ('telefono_contacto',  'VARCHAR(32)'),
        ('fecha_atencion',     'DATE'),
    ]
    for col, tipo in cols_tipos_migrantes:
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE migrantes MODIFY COLUMN {col} {tipo}")
            con.commit()
            cur.close()
        except Exception as e:
            print(f'[migración tipo {col}] {e}')

    # Migración: eliminar tabla dependientes_migrante (ya no se usa).
    try:
        cur = con.cursor()
        cur.execute("DROP TABLE IF EXISTS dependientes_migrante")
        con.commit()
        cur.close()
    except Exception as e:
        pass

    # Migración: tabla solicitudes_arco_rect para Derechos ARCO.
    try:
        cur = con.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS solicitudes_arco_rect ("
            "  id              INT AUTO_INCREMENT PRIMARY KEY,"
            "  migrante_id     INT NOT NULL,"
            "  solicitado_por  VARCHAR(64) NOT NULL,"
            "  cambios_json    TEXT NOT NULL,"
            "  estado          ENUM('pendiente','aprobada','rechazada') DEFAULT 'pendiente',"
            "  motivo_rechazo  TEXT,"
            "  fecha_solicitud DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,"
            "  resuelto_por    VARCHAR(64),"
            "  resuelto_en     DATETIME,"
            "  FOREIGN KEY (migrante_id) REFERENCES migrantes(id) ON DELETE CASCADE,"
            "  INDEX idx_rect_estado (estado)"
            ") ENGINE=InnoDB"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración solicitudes_arco_rect] {e}')

    # Migración: ampliar ENUM de log para eventos ARCO y registro workflow.
    try:
        cur = con.cursor()
        cur.execute(
            "ALTER TABLE log_certificados MODIFY resultado "
            "ENUM('login_exitoso','login_fallido','pwd_incorrecta',"
            "     'cert_revocado','cert_expirado','cert_emitido',"
            "     'cert_revocado_admin','rol_modificado','pwd_cambiada',"
            "     'acceso_denegado','migrante_registrado','migrante_actualizado',"
            "     'migrante_solicitud_eliminacion','migrante_eliminado',"
            "     'pwd_reset_solicitado','pwd_reset_aprobado','pwd_reset_rechazado',"
            "     'arco_acceso','arco_rect_solicitada','arco_rect_aprobada',"
            "     'arco_rect_rechazada','arco_cancel_solicitada',"
            "     'arco_cancel_op_solicitada','arco_cancel_coord_firmada',"
            "     'arco_cancel_coord_rechazada',"
            "     'registro_mig_enviado_voluntario','registro_mig_enviado_op_directo',"
            "     'registro_mig_op_validado','registro_mig_op_rechazado',"
            "     'registro_mig_coord_aprobado','registro_mig_coord_rechazado',"
            "     'registro_mig_admin_aprobado','registro_mig_insertado') NOT NULL"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración log ENUM ARCO] {e}')

    # Migración: claves EC para coordinadores.
    for col in ('ec_private_key', 'ec_public_key'):
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE usuarios ADD COLUMN {col} TEXT")
            con.commit()
            cur.close()
        except Exception:
            pass

    # Migración: tabla solicitudes_cancelacion_op (op → coord).
    try:
        cur = con.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS solicitudes_cancelacion_op ("
            "  id              INT AUTO_INCREMENT PRIMARY KEY,"
            "  migrante_id     INT NOT NULL,"
            "  solicitado_por  VARCHAR(120) NOT NULL,"
            "  motivo          TEXT,"
            "  estado          ENUM('pendiente','aprobada','rechazada') DEFAULT 'pendiente',"
            "  resuelto_por    VARCHAR(120),"
            "  resuelto_en     DATETIME,"
            "  fecha_solicitud DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,"
            "  FOREIGN KEY (migrante_id) REFERENCES migrantes(id) ON DELETE CASCADE"
            ") ENGINE=InnoDB"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración solicitudes_cancelacion_op] {e}')

    # Migración: campos EC en solicitudes_eliminacion.
    for col in ('solicitante_op VARCHAR(120)', 'firma_coord TEXT',
                'coord_pubkey TEXT', 'mensaje_firmado TEXT'):
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE solicitudes_eliminacion ADD COLUMN {col}")
            con.commit()
            cur.close()
        except Exception:
            pass

    # Generar claves EC para coordinadores existentes sin clave.
    try:
        cur = con.cursor()
        cur.execute(
            "SELECT usuario FROM usuarios WHERE rol='coord'"
            " AND (ec_private_key IS NULL OR ec_private_key='')"
        )
        coords_sin_clave = cur.fetchall()
        cur.close()
        for row in coords_sin_clave:
            priv_pem, pub_pem = generar_claves_ec()
            cur = con.cursor()
            cur.execute(
                "UPDATE usuarios SET ec_private_key=%s, ec_public_key=%s WHERE usuario=%s",
                (priv_pem, pub_pem, row['usuario'])
            )
            con.commit()
            cur.close()
    except Exception as e:
        print(f'[migración claves EC coord] {e}')

    # Generar claves EC para operativos existentes sin clave.
    try:
        cur = con.cursor()
        cur.execute(
            "SELECT usuario FROM usuarios WHERE rol='op'"
            " AND (ec_private_key IS NULL OR ec_private_key='')"
        )
        ops_sin_clave = cur.fetchall()
        cur.close()
        for row in ops_sin_clave:
            priv_pem, pub_pem = generar_claves_ec()
            cur = con.cursor()
            cur.execute(
                "UPDATE usuarios SET ec_private_key=%s, ec_public_key=%s WHERE usuario=%s",
                (priv_pem, pub_pem, row['usuario'])
            )
            con.commit()
            cur.close()
    except Exception as e:
        print(f'[migración claves EC op] {e}')

    # Migración: tabla solicitudes_registro_migrante (workflow de aprobación).
    try:
        cur = con.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS solicitudes_registro_migrante ("
            "  id                    INT AUTO_INCREMENT PRIMARY KEY,"
            "  nombre                VARCHAR(200) NOT NULL,"
            "  fecha_nacimiento      DATE,"
            "  pais_origen           VARCHAR(80),"
            "  fecha_atencion        DATE,"
            "  genero                VARCHAR(30),"
            "  departamento_estado   VARCHAR(100),"
            "  estado_civil          VARCHAR(30),"
            "  grupo_poblacion       VARCHAR(80),"
            "  telefono_contacto     VARCHAR(32),"
            "  estado                ENUM('pendiente_op','pendiente_coord','aprobada','rechazada')"
            "                        NOT NULL DEFAULT 'pendiente_op',"
            "  enviado_por           VARCHAR(120) NOT NULL,"
            "  rol_enviado_por       ENUM('voluntario','op') NOT NULL,"
            "  op_validador          VARCHAR(120),"
            "  firma_op              TEXT,"
            "  op_pubkey             TEXT,"
            "  mensaje_firmado_op    TEXT,"
            "  op_validado_en        DATETIME,"
            "  coord_aprobador       VARCHAR(120),"
            "  firma_coord           TEXT,"
            "  coord_pubkey          TEXT,"
            "  mensaje_firmado_coord TEXT,"
            "  coord_aprobado_en     DATETIME,"
            "  rechazado_por         VARCHAR(120),"
            "  motivo_rechazo        TEXT,"
            "  rechazado_en          DATETIME,"
            "  creado                DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,"
            "  actualizado           DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP"
            "                        ON UPDATE CURRENT_TIMESTAMP,"
            "  INDEX idx_srm_estado  (estado),"
            "  INDEX idx_srm_enviado (enviado_por)"
            ") ENGINE=InnoDB"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración solicitudes_registro_migrante] {e}')

    # Migración: columna firmas_aprobacion en migrantes.
    try:
        cur = con.cursor()
        cur.execute("ALTER TABLE migrantes ADD COLUMN firmas_aprobacion TEXT")
        con.commit()
        cur.close()
    except Exception:
        pass

    # Migración: columnas de override admin en solicitudes_registro_migrante.
    for col_def in (
        'admin_aprobador      VARCHAR(100)',
        'admin_aprobado_en    DATETIME',
        'firma_admin          TEXT',
        'admin_pubkey         TEXT',
        'mensaje_firmado_admin TEXT',
    ):
        try:
            cur = con.cursor()
            cur.execute(f'ALTER TABLE solicitudes_registro_migrante ADD COLUMN {col_def}')
            con.commit()
            cur.close()
        except Exception:
            pass

    # Migración: tabla historial_eliminaciones (archivo permanente de migrantes eliminados).
    try:
        cur = con.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS historial_eliminaciones ("
            "  id                INT AUTO_INCREMENT PRIMARY KEY,"
            "  folio             VARCHAR(20),"
            "  nombre_migrante   VARCHAR(200),"
            "  fecha_solicitud   DATETIME,"
            "  solicitado_por    VARCHAR(120),"
            "  motivo            TEXT,"
            "  firma_coord       TEXT,"
            "  coord_pubkey      TEXT,"
            "  mensaje_firmado   TEXT,"
            "  aprobado_por      VARCHAR(120),"
            "  tipo              VARCHAR(20) NOT NULL DEFAULT 'directa',"
            "  es_arco           TINYINT(1) NOT NULL DEFAULT 0,"
            "  fecha_eliminacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,"
            "  INDEX idx_hist_fecha (fecha_eliminacion)"
            ") ENGINE=InnoDB"
        )
        con.commit()
        cur.close()
    except Exception as e:
        print(f'[migración historial_eliminaciones] {e}')

    # Migración: columna es_arco en solicitudes_eliminacion e historial_eliminaciones.
    for tabla in ('solicitudes_eliminacion', 'historial_eliminaciones'):
        try:
            cur = con.cursor()
            cur.execute(
                f"ALTER TABLE {tabla} ADD COLUMN es_arco TINYINT(1) NOT NULL DEFAULT 0"
            )
            con.commit()
            cur.close()
        except Exception:
            pass

    # Migración: columnas de bloqueo por intentos fallidos.
    for col_def in (
        'intentos_fallidos INT NOT NULL DEFAULT 0',
        'bloqueado TINYINT(1) NOT NULL DEFAULT 0',
    ):
        try:
            cur = con.cursor()
            cur.execute(f"ALTER TABLE usuarios ADD COLUMN {col_def}")
            con.commit()
            cur.close()
        except Exception:
            pass

    # Sembrar solo si la tabla está vacía
    cur = con.cursor()
    cur.execute('SELECT COUNT(*) c FROM usuarios')
    if cur.fetchone()['c'] == 0:
        for login_, u in USUARIOS_DEMO.items():
            cur.execute(
                'INSERT INTO usuarios (usuario, nombre, rol, password_hash, debe_cambiar_pwd, activo)'
                ' VALUES (%s,%s,%s,%s,0,%s)',
                (login_, u['nombre'], u['rol'], generate_password_hash(u['password']), u.get('activo', 1))
            )
        for c in CERTIFICADOS_SEED:
            estado = 'vigente' if c['activo'] else 'revocado'
            cur.execute(
                'INSERT INTO certificados (serial, usuario, rol, fecha_emision, estado)'
                ' VALUES (%s,%s,%s,%s,%s)',
                (c['serial'], c['usuario'], c['rol'], c['fecha_emision'], estado)
            )
            if not c['activo']:
                cur.execute(
                    'INSERT IGNORE INTO certificados_revocados (serial, usuario, rol, motivo)'
                    ' VALUES (%s,%s,%s,%s)',
                    (c['serial'], c['usuario'], c['rol'], 'Seed inicial')
                )
        con.commit()
    cur.close()
    con.close()

# ── HELPERS ───────────────────────────────────────────────────
def generar_contrasena_temporal(longitud=14):
    especiales = '!@#$%&*'
    chars = string.ascii_letters + string.digits + especiales
    while True:
        pwd = ''.join(secrets.choice(chars) for _ in range(longitud))
        if (any(c.isupper() for c in pwd) and
                any(c.isdigit() for c in pwd) and
                any(c in especiales for c in pwd)):
            return pwd

def validar_contrasena(pwd: str):
    if len(pwd) < 12:
        return False, 'La contraseña debe tener al menos 12 caracteres.'
    if not re.search(r'[A-Z]', pwd):
        return False, 'La contraseña debe contener al menos una letra mayúscula.'
    if not re.search(r'\d', pwd):
        return False, 'La contraseña debe contener al menos un número.'
    if not re.search(r'[^A-Za-z0-9]', pwd):
        return False, 'La contraseña debe contener al menos un carácter especial.'
    return True, ''

def sanitizar_nombre(nombre):
    return re.sub(r'[^a-zA-ZáéíóúÁÉÍÓÚñÑüÜ0-9 \-]', '', nombre).strip()

def validar_email(email):
    """Valida formato básico de email"""
    import re
    patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(patron, email) is not None

def extraer_email_de_cert(cert):
    """Extrae el email del certificado X.509"""
    try:
        # Intenta obtener del campo SubjectAlternativeName
        san_ext = cert.extensions.get_extension_for_oid(
            x509.oid.ExtensionOID.SUBJECT_ALTERNATIVE_NAME
        )
        for name in san_ext.value:
            if isinstance(name, x509.RFC822Name):
                return name.value
    except x509.ExtensionNotFound:
        pass
    
    # Fallback: buscar en Subject
    try:
        email_attrs = cert.subject.get_attributes_for_oid(NameOID.EMAIL_ADDRESS)
        if email_attrs:
            return email_attrs[0].value
    except Exception:
        pass
    
    return None
    
def leer_password_admin():
    with open("admin_bootstrap.txt", "r") as f:
        for linea in f:
            if "Password:" in linea:
                return linea.split("Password:")[1].strip()
    return None


def leer_certificado_usb(password: str, email: str):
    """
    Detecta un certificado en USB simulada y valida que corresponda al email del usuario.
    Args:
        password: contraseña para desbloquear el .p12
        email: email del usuario que intenta entrar
    Returns:
        certificate (objeto x509) o None si no se encontró/validó
    """
    rutas = glob.glob("usb_simulada/*.p12")
    if not rutas:
        return None

    for ruta in rutas:
        try:
            with open(ruta, "rb") as f:
                data = f.read()
            private_key, certificate, _ = pkcs12.load_key_and_certificates(
                data, password.encode("utf-8")
            )
            if certificate:
                cert_email = extraer_email_de_cert(certificate)
                if cert_email and cert_email.lower() == email.lower():
                    return certificate
        except Exception as e:
            print(f"[USB] Error al leer {ruta}: {e}")
            continue

    return None
       
def generar_p12_real(email, nombre, rol, password):
    """
    Genera certificado P12 con email incluido.
    
    Args:
        email: Email del usuario (ej: juan.perez@casamonarca.org)
        nombre: Nombre completo (ej: Juan Pérez)
        rol: Rol del usuario (admin/coord/op)
        password: Contraseña para proteger el .p12
    
    Returns:
        (p12_bytes, serial_hex)
    """
    # Cargar la CA para firmar
    ca_cert, ca_key = cargar_ca()

    private_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=2048,
    )

    # Subject con email
    subject = x509.Name([
        x509.NameAttribute(NameOID.COUNTRY_NAME, "MX"),
        x509.NameAttribute(NameOID.ORGANIZATION_NAME, "Casa Monarca"),
        x509.NameAttribute(NameOID.ORGANIZATIONAL_UNIT_NAME, rol),
        x509.NameAttribute(NameOID.COMMON_NAME, nombre),
        x509.NameAttribute(NameOID.EMAIL_ADDRESS, email),  # ← NUEVO
    ])

    # Crear certificado
    cert = (
        x509.CertificateBuilder()
        .subject_name(subject)
        .issuer_name(ca_cert.subject)
        .public_key(private_key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(datetime.utcnow() - timedelta(minutes=1))
        .not_valid_after(datetime.utcnow() + timedelta(days=30))
        .add_extension(
            x509.BasicConstraints(ca=False, path_length=None),
            critical=True
        )
        .add_extension(
            # ← NUEVO: Email en SubjectAlternativeName
            x509.SubjectAlternativeName([
                x509.RFC822Name(email)
            ]),
            critical=False
        )
        .sign(ca_key, hashes.SHA256())
    )

    p12_bytes = pkcs12.serialize_key_and_certificates(
        name=nombre.encode("utf-8"),
        key=private_key,
        cert=cert,
        cas=[ca_cert],
        encryption_algorithm=serialization.BestAvailableEncryption(
            password.encode("utf-8")
        )
    )

    return p12_bytes, format(cert.serial_number, 'X').upper(), cert

def badge_rol(rol):
    if rol == 'admin':
        return 'Administrador'
    if rol == 'coord':
        return 'Coordinador'
    if rol == 'voluntario':
        return 'Voluntario'
    return 'Operador'

def url_panel(rol):
    """Devuelve la URL del panel correcto según el rol."""
    if rol == 'admin':
        return url_for('panel_admin')
    if rol == 'voluntario':
        return url_for('panel_voluntario')
    return url_for('panel_usuario')

def log_evento(resultado, usuario=None, rol=None, serial=None, detalle=None):
    """Registra evento en la tabla de auditoría.
    Pedido explícito del módulo de Santiago (sección 3.7 del reporte)."""
    try:
        db = obtener_db()
        _exec(db,
            'INSERT INTO log_certificados'
            ' (serial, usuario, rol, ip_origen, resultado, detalle)'
            ' VALUES (%s,%s,%s,%s,%s,%s)',
            (serial, usuario, rol, request.remote_addr, resultado, detalle)
        )
        db.commit()
    except Exception as e:
        print(f'[log_evento] error: {e}')

def refrescar_expirados():
    """Marca como 'expirado' todo cert cuya fecha ya pasó."""
    db = obtener_db()
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _exec(db,
        "UPDATE certificados SET estado='expirado'"
        " WHERE estado='vigente' AND fecha_expiracion IS NOT NULL"
        " AND fecha_expiracion < %s",
        (ahora,)
    )
    db.commit()

# ── CONFIGURACIÓN DE mTLS ─────────────────────────────────────
# MTLS_ENABLED=1 → el servidor exige cert de cliente (modo producción / demo).
# MTLS_ENABLED=0 → arranca en HTTP plano (modo debug). Default: 1.
MTLS_ENABLED = os.environ.get('MTLS_ENABLED', '1') == '1'

# Rutas públicas que NO requieren cert de cliente
# (ninguna en este diseño: el cert se pide para TODO el sitio)
RUTAS_SIN_MTLS = set()

@app.before_request
def verificar_cert_cliente():
    """
    Middleware mTLS (corregido):
    Antes de CUALQUIER ruta, el servidor ya NO exige que el navegador
    presente un certificado de cliente. Esto permite que usuarios normales
    entren solo con usuario/contraseña.

    La validación de certificados emitidos por la CA de Casa Monarca
    se mantiene en el flujo USB para roles admin/coord.
    """
    # Si mTLS está desactivado, no hacer nada
    if not MTLS_ENABLED:
        return None

    # 🔴 Cambio: ya no exigimos certificado de cliente en el handshake TLS
    # Se deja pasar a todas las rutas, y la validación fuerte se hace
    # en login_usb para roles altos.
    return None

# ── DECORADOR DE AUTENTICACIÓN ────────────────────────────────
def verificar_certificado(f):
    @wraps(f)
    def decorado(*args, **kwargs):
        if 'rol' not in session:
            return redirect(url_for('login'))
        g.rol = session['rol']
        g.usuario = session.get('usuario', '')
        g.nombre = session.get('nombre', 'Usuario')
        return f(*args, **kwargs)
    return decorado

# ── RUTAS DE AUTENTICACIÓN ────────────────────────────────────
@app.route('/')
def index():
    if 'rol' not in session:
        return redirect(url_for('login'))
    return redirect(url_panel(session['rol']))

@app.route('/login', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    # POST: verificar credenciales
    email = request.form.get('usuario', '').strip().lower()
    password = request.form.get('password', '')
    if not email or not password:
        return render_template('login.html',
                               error='Email y contraseña son requeridos')

    # Validar formato de email
    if not validar_email(email):
        return render_template('login.html',
                               error='Formato de email inválido')

    db = obtener_db()
    usuario = _exec(db,
        'SELECT * FROM usuarios WHERE usuario=%s', (email,)
    ).fetchone()

    # Verificar cuenta bloqueada
    if usuario and usuario.get('bloqueado', 0):
        log_evento('acceso_denegado', usuario=email,
                   detalle='Intento de acceso a cuenta bloqueada')
        return render_template('login.html',
                               error='Cuenta bloqueada por múltiples intentos fallidos. Contacta al administrador.')

    # Verificar credenciales
    if not usuario or not check_password_hash(usuario['password_hash'], password):
        if usuario:
            try:
                _exec(db,
                    'UPDATE usuarios SET intentos_fallidos = COALESCE(intentos_fallidos, 0) + 1,'
                    ' bloqueado = CASE WHEN COALESCE(intentos_fallidos, 0) + 1 >= 10 THEN 1 ELSE COALESCE(bloqueado, 0) END'
                    ' WHERE usuario=%s', (email,))
                db.commit()
            except Exception:
                pass
        log_evento('login_fallido', usuario=email,
                   detalle='Credenciales incorrectas')
        return render_template('login.html',
                               error='Email o contraseña incorrectos')

    # Verificar que la cuenta esté activa
    if not usuario.get('activo', 1):
        log_evento('acceso_denegado', usuario=email,
                   detalle='Intento de acceso a cuenta inactiva')
        return render_template('login.html',
                               error='Tu cuenta está inactiva. Contacta al administrador.')

    # Resetear contador de intentos en login exitoso (antes de continuar)
    try:
        _exec(db, 'UPDATE usuarios SET intentos_fallidos=0, bloqueado=0 WHERE usuario=%s', (email,))
        db.commit()
    except Exception:
        pass

    # Si el rol es admin o coord, redirigir a login_usb
    if usuario['rol'] in ('admin', 'coord'):
        session.clear()
        session['pending_user'] = usuario['usuario']
        session['nombre'] = usuario['nombre']
        session['rol'] = usuario['rol']
        return redirect(url_for('login_usb'))

    # Validación 1: Estado del certificado en BD
    refrescar_expirados()

    cert = _exec(db,
        'SELECT estado, serial FROM certificados WHERE usuario=%s OR usuario=%s'
        ' ORDER BY id DESC LIMIT 1',
        (email, usuario['nombre'])
    ).fetchone()

    if cert and cert['estado'] == 'revocado':
        log_evento('cert_revocado', usuario=email, rol=usuario['rol'],
                   serial=cert['serial'])
        return render_template('login.html',
                               error='Tu certificado fue revocado. Contacta al administrador.')

    if cert and cert['estado'] == 'expirado':
        log_evento('cert_expirado', usuario=email, rol=usuario['rol'],
                   serial=cert['serial'])
        return render_template('login.html',
                               error='Tu certificado expiró. Solicita uno nuevo.')

    # Login exitoso - Crear sesión
    session.clear()
    session['usuario'] = usuario['usuario']  # email
    session['nombre'] = usuario['nombre']
    session['rol'] = usuario['rol']

    log_evento('login_exitoso', usuario=email, rol=usuario['rol'],
               serial=cert['serial'] if cert else None)

    if usuario['debe_cambiar_pwd']:
        return redirect(url_for('cambiar_pwd'))

    return redirect(url_panel(usuario['rol']))
    
@app.route('/login_usb', methods=['GET', 'POST'])
def login_usb():
    # Mostrar formulario para subir el .p12
    if request.method == 'GET':
        return render_template('login_usb.html')

    # Procesar el archivo .p12
    archivo = request.files.get('certificado_usb')
    if not archivo:
        log_evento('acceso_denegado', usuario=session.get('pending_user'),
                   detalle='No se seleccionó archivo USB')
        return render_template('login_usb.html',
                               error='Debes seleccionar tu archivo .p12 de la USB')

    p12_password = request.form.get('p12_password', '')
    email = session.get('pending_user')

    try:
        data = archivo.read()
        private_key, certificate, _ = pkcs12.load_key_and_certificates(
            data, p12_password.encode("utf-8")
        )
    except Exception as e:
        log_evento('acceso_denegado', usuario=email,
                   detalle=f'Error al abrir certificado USB: {str(e)}')
        return render_template('login_usb.html',
                               error='No se pudo abrir el certificado USB')

    cert_email = extraer_email_de_cert(certificate)
    if not cert_email or cert_email.lower() != email.lower():
        log_evento('acceso_denegado', usuario=email,
                   detalle=f'Cert USB email={cert_email}, login email={email}')
        return render_template('login_usb.html',
                               error='El certificado USB no corresponde a este usuario')

    # Validación en BD
    db = obtener_db()
    refrescar_expirados()
    cert = _exec(db,
        'SELECT estado, serial FROM certificados WHERE usuario=%s OR usuario=%s '
        'ORDER BY id DESC LIMIT 1',
        (email, session.get('nombre'))
    ).fetchone()

    if cert and cert['estado'] == 'revocado':
        log_evento('cert_revocado', usuario=email, rol=session['rol'],
                   serial=cert['serial'])
        return render_template('login_usb.html',
                               error='Tu certificado fue revocado. Contacta al administrador.')
    if cert and cert['estado'] == 'expirado':
        log_evento('cert_expirado', usuario=email, rol=session['rol'],
                   serial=cert['serial'])
        return render_template('login_usb.html',
                               error='Tu certificado expiró. Solicita uno nuevo.')

    # Login exitoso: completar sesión
    session['usuario'] = email
    log_evento('login_exitoso', usuario=email, rol=session['rol'],
               serial=cert['serial'] if cert else None)

    db = obtener_db()
    usuario = _exec(db, 'SELECT * FROM usuarios WHERE usuario=%s', (email,)).fetchone()
    if usuario and usuario['debe_cambiar_pwd']:
        return redirect(url_for('cambiar_pwd'))

    return redirect(url_panel(session['rol']))

@app.route('/cambiar_pwd', methods=['GET', 'POST'])
@verificar_certificado
def cambiar_pwd():
    error = None
    if request.method == 'POST':
        nueva  = request.form.get('nueva', '')
        confir = request.form.get('confirmacion', '')
        valida, msg_val = validar_contrasena(nueva)
        if not valida:
            error = msg_val
        elif nueva != confir:
            error = 'Las contraseñas no coinciden.'
        else:
            db = obtener_db()
            fila = _exec(db, 'SELECT password_hash FROM usuarios WHERE usuario=%s',
                         (g.usuario,)).fetchone()
            if fila and check_password_hash(fila['password_hash'], nueva):
                error = 'La nueva contraseña no puede ser igual a la contraseña actual.'
            else:
                _exec(db,
                    'UPDATE usuarios SET password_hash=%s, debe_cambiar_pwd=0'
                    ' WHERE usuario=%s',
                    (generate_password_hash(nueva), g.usuario)
                )
                db.commit()
                log_evento('pwd_cambiada', usuario=g.usuario, rol=g.rol)
                return redirect(url_panel(g.rol))
    return render_template('cambiar_pwd.html', error=error, nombre=g.nombre)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── PANEL PRINCIPAL ───────────────────────────────────────────
@app.route('/admin')
@verificar_certificado
def panel_admin():
    print(">>> Renderizando panel_admin")
    if g.rol != 'admin':
        abort(403)
    return render_template('admin.html', nombre=g.nombre, rol=g.rol, usuario=g.usuario)

# ── API: MÉTRICAS ─────────────────────────────────────────────
@app.route('/admin/api/metricas')
@verificar_certificado
def api_metricas():
    if g.rol != 'admin':
        abort(403)
    refrescar_expirados()
    db = obtener_db()
    total     = _exec(db, 'SELECT COUNT(*) c FROM certificados').fetchone()['c']
    vigentes  = _exec(db, "SELECT COUNT(*) c FROM certificados WHERE estado='vigente'").fetchone()['c']
    revocados = _exec(db, "SELECT COUNT(*) c FROM certificados WHERE estado='revocado'").fetchone()['c']
    expirados = _exec(db, "SELECT COUNT(*) c FROM certificados WHERE estado='expirado'").fetchone()['c']
    # 'activos' se mantiene como alias para no romper el front existente.
    return jsonify(total=total, activos=vigentes, vigentes=vigentes,
                   revocados=revocados, expirados=expirados)

# ── API: ESTADÍSTICAS PANTALLA DE INICIO ─────────────────────
@app.route('/admin/api/inicio_stats')
@verificar_certificado
def api_inicio_stats():
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()

    # ── Colaboradores ─────────────────────────────────────────
    roles_rows = _exec(db,
        "SELECT rol, COUNT(*) c FROM usuarios WHERE activo=1 GROUP BY rol"
    ).fetchall()
    roles = {r['rol']: r['c'] for r in roles_rows}
    colab = {
        'total':       sum(roles.values()),
        'coordinadores': roles.get('coord', 0),
        'operativos':  roles.get('op', 0),
        'voluntarios': roles.get('voluntario', 0),
    }

    # ── Operaciones pendientes ────────────────────────────────
    altas = _exec(db,
        "SELECT COUNT(*) c FROM solicitudes_registro_migrante"
        " WHERE estado IN ('pendiente_op','pendiente_coord')"
    ).fetchone()['c']
    rect = _exec(db,
        "SELECT COUNT(*) c FROM solicitudes_arco_rect WHERE estado='pendiente'"
    ).fetchone()['c']
    elim = _exec(db,
        "SELECT COUNT(*) c FROM solicitudes_eliminacion WHERE estado='pendiente' AND es_arco=0"
    ).fetchone()['c']
    arco = _exec(db,
        "SELECT COUNT(*) c FROM solicitudes_eliminacion WHERE estado='pendiente' AND es_arco=1"
    ).fetchone()['c']
    pendientes = {
        'altas_migrantes':    altas,
        'rectificaciones':    rect,
        'eliminacion':        elim,
        'procesos_arco':      arco,
    }

    # ── Migrantes ─────────────────────────────────────────────
    total_mig = _exec(db, 'SELECT COUNT(*) c FROM migrantes').fetchone()['c']

    paises_rows = _exec(db,
        'SELECT pais_origen, COUNT(*) c FROM migrantes'
        ' WHERE pais_origen IS NOT NULL AND pais_origen != ""'
        ' GROUP BY pais_origen ORDER BY c DESC LIMIT 3'
    ).fetchall()
    top_paises = [{'pais': r['pais_origen'], 'cnt': r['c']} for r in paises_rows]

    genero_rows = _exec(db,
        "SELECT genero, COUNT(*) c,"
        "  ROUND(AVG(TIMESTAMPDIFF(YEAR, fecha_nacimiento, CURDATE())), 1) AS edad_media"
        " FROM migrantes WHERE genero IS NOT NULL AND fecha_nacimiento IS NOT NULL"
        " GROUP BY genero"
    ).fetchall()
    generos = {}
    for r in genero_rows:
        generos[r['genero']] = {'cnt': r['c'], 'edad_media': float(r['edad_media']) if r['edad_media'] else None}

    ninos_row = _exec(db,
        "SELECT COUNT(*) c,"
        "  ROUND(AVG(TIMESTAMPDIFF(YEAR, fecha_nacimiento, CURDATE())), 1) AS edad_media"
        " FROM migrantes"
        " WHERE fecha_nacimiento IS NOT NULL"
        "   AND TIMESTAMPDIFF(YEAR, fecha_nacimiento, CURDATE()) < 18"
    ).fetchone()
    ninos = {
        'cnt': ninos_row['c'] if ninos_row else 0,
        'edad_media': float(ninos_row['edad_media']) if ninos_row and ninos_row['edad_media'] else None,
    }

    migrantes = {
        'total':      total_mig,
        'top_paises': top_paises,
        'generos':    generos,
        'ninos':      ninos,
    }

    return jsonify(colaboradores=colab, pendientes=pendientes, migrantes=migrantes)


# ── API: LISTA DE CERTIFICADOS ────────────────────────────────
@app.route('/admin/api/certificados')
@verificar_certificado
def api_certificados():
    if g.rol != 'admin':
        abort(403)

    refrescar_expirados()
    db = obtener_db()
    rows = _exec(db,
        'SELECT id, serial, usuario, rol, fecha_emision, fecha_expiracion, estado'
        ' FROM certificados ORDER BY id DESC'
    ).fetchall()

    certificados_ui = []
    ahora = datetime.now()
    for c in rows:
        dias = None
        fe = c['fecha_expiracion']
        if fe:
            if hasattr(fe, 'replace'):
                fe_dt = fe
            else:
                try:
                    fe_dt = datetime.strptime(str(fe)[:19], '%Y-%m-%d %H:%M:%S')
                except Exception:
                    fe_dt = None
            if fe_dt:
                dias = (fe_dt - ahora).days
        certificados_ui.append({
            'id': c['id'],
            'nombre': c['usuario'],
            'usuario': c['usuario'],
            'rol': c['rol'],
            'serial': c['serial'],
            'fecha': c['fecha_emision'],
            'fecha_emision': c['fecha_emision'],
            'fecha_expiracion': str(c['fecha_expiracion'])[:10] if c['fecha_expiracion'] else None,
            'estado': c['estado'],
            'activo': c['estado'] == 'vigente',
            'dias_para_vencer': dias,
        })

    return jsonify(certificados_ui)

# ── API: LOG DE AUDITORÍA ─────────────────────────────────────
@app.route('/admin/api/log')
@verificar_certificado
def api_log():
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    try:
        rows = _exec(db,
            'SELECT fecha, usuario, rol, ip_origen, resultado, detalle'
            ' FROM log_certificados ORDER BY id DESC LIMIT 50'
        ).fetchall()
        ARCO_RESULTADOS = {
            'arco_acceso', 'arco_rect_solicitada', 'arco_rect_aprobada',
            'arco_rect_rechazada', 'arco_cancel_solicitada',
            'arco_cancel_op_solicitada', 'arco_cancel_coord_firmada',
            'arco_cancel_coord_rechazada',
        }
        result = []
        for r in rows:
            d = dict(r)
            d['es_arco'] = d.get('resultado', '') in ARCO_RESULTADOS
            result.append(d)
        return jsonify(result)
    
        
    except Exception as e:
        print(f"[ERROR api_log] {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/api/eliminados')
@verificar_certificado
def api_eliminados():
    if g.rol != 'admin':
        abort(403)
    fecha_inicio = request.args.get('fecha_inicio', '').strip() or None
    fecha_fin    = request.args.get('fecha_fin',    '').strip() or None
    sql    = 'SELECT * FROM historial_eliminaciones WHERE 1=1'
    params = []
    if fecha_inicio:
        sql += ' AND fecha_eliminacion >= %s'; params.append(fecha_inicio)
    if fecha_fin:
        sql += ' AND fecha_eliminacion <= %s'; params.append(fecha_fin + ' 23:59:59')
    sql += ' ORDER BY id DESC LIMIT 500'
    db   = obtener_db()
    rows = _exec(db, sql, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in ('fecha_solicitud', 'fecha_eliminacion'):
            if d.get(k) and hasattr(d[k], 'isoformat'):
                d[k] = d[k].isoformat()
        result.append(d)
    return jsonify(result)


@app.route('/admin/api/eliminados/<int:hid>/firma')
@verificar_certificado
def api_eliminados_firma(hid):
    if g.rol != 'admin':
        abort(403)
    db  = obtener_db()
    row = _exec(db, 'SELECT * FROM historial_eliminaciones WHERE id=%s', (hid,)).fetchone()
    if not row:
        return jsonify(error='Registro no encontrado'), 404
    r = dict(row)
    for k in ('fecha_solicitud', 'fecha_eliminacion'):
        if r.get(k) and hasattr(r[k], 'isoformat'):
            r[k] = r[k].isoformat()
    if r.get('firma_coord') and r.get('coord_pubkey') and r.get('mensaje_firmado'):
        r['firma_valida'] = verificar_firma(r['coord_pubkey'], r['mensaje_firmado'], r['firma_coord'])
    else:
        r['firma_valida'] = None
    return jsonify(r)


# ── DESCARGAS EXCEL ───────────────────────────────────────────

def _make_excel(headers, rows):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    hf  = PatternFill("solid", fgColor="E62055")
    hft = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hft
        cell.alignment = Alignment(horizontal="center")
    for row_data in rows:
        ws.append(list(row_data))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route('/admin/descargar/migrantes')
@verificar_certificado
def descargar_excel_migrantes():
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    fecha_inicio    = request.args.get('fecha_inicio', '').strip() or None
    fecha_fin       = request.args.get('fecha_fin',    '').strip() or None
    pais_origen     = request.args.get('pais_origen',  '').strip() or None
    grupo_poblacion = request.args.get('grupo_poblacion', '').strip() or None
    if fecha_inicio and not fecha_fin:
        from datetime import date as _date_hoy
        fecha_fin = _date_hoy.today().strftime('%Y-%m-%d')

    sql    = ('SELECT folio, nombre, fecha_nacimiento, pais_origen, grupo_poblacion, '
              'fecha_atencion, genero, departamento_estado, estado_civil, '
              'telefono_contacto, registrado_por, firmas_aprobacion, creado '
              'FROM migrantes WHERE 1=1')
    params = []
    if fecha_inicio:
        sql += ' AND fecha_atencion >= %s'; params.append(fecha_inicio)
    if fecha_fin:
        sql += ' AND fecha_atencion <= %s'; params.append(fecha_fin)
    if pais_origen:
        sql += ' AND pais_origen LIKE %s'; params.append(f'%{pais_origen}%')
    if grupo_poblacion:
        sql += ' AND grupo_poblacion LIKE %s'; params.append(f'%{grupo_poblacion}%')
    sql += ' ORDER BY id DESC'

    db   = obtener_db()
    rows = _exec(db, sql, params).fetchall()

    import json as _json
    def _s(v): return str(v)[:10] if v else ''
    def _dt(v): return str(v)[:19] if v else ''

    from datetime import date as _date
    def _edad(fnac):
        if not fnac:
            return ''
        hoy = _date.today()
        try:
            if hasattr(fnac, 'year'):
                d = fnac
            else:
                from datetime import datetime
                d = datetime.strptime(str(fnac)[:10], '%Y-%m-%d').date()
            return hoy.year - d.year - ((hoy.month, hoy.day) < (d.month, d.day))
        except Exception:
            return ''

    def _firmantes(json_str):
        if not json_str:
            return '', ''
        try:
            firmas = _json.loads(json_str)
        except Exception:
            return '', ''
        op    = (firmas.get('op') or {}).get('usuario', '')
        coord = (firmas.get('coord') or firmas.get('admin') or {}).get('usuario', '')
        return op, coord

    headers = ['Folio', 'Nombre', 'Fecha de nacimiento', 'Edad',
               'País de origen', 'Grupo de población',
               'Fecha de atención', 'Género', 'Departamento/Estado',
               'Estado civil', 'Teléfono', 'Registrado por',
               'Firmado por (operativo)', 'Firmado por (coordinador/admin)',
               'Fecha de creación']
    data = []
    for r in rows:
        op_firma, coord_firma = _firmantes(r['firmas_aprobacion'])
        data.append((
            r['folio'], r['nombre'],
            _s(r['fecha_nacimiento']), _edad(r['fecha_nacimiento']),
            r['pais_origen'], r['grupo_poblacion'],
            _s(r['fecha_atencion']), r['genero'], r['departamento_estado'],
            r['estado_civil'], r['telefono_contacto'], r['registrado_por'],
            op_firma, coord_firma,
            _dt(r['creado']),
        ))
    buf = _make_excel(headers, data)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='migrantes.xlsx')


@app.route('/admin/descargar/migrante/<int:mid>')
@verificar_certificado
def descargar_excel_migrante_individual(mid):
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    import json as _json
    r = _exec(db,
        'SELECT folio, nombre, fecha_nacimiento, pais_origen, grupo_poblacion,'
        ' fecha_atencion, genero, departamento_estado, estado_civil,'
        ' telefono_contacto, registrado_por, firmas_aprobacion, creado'
        ' FROM migrantes WHERE id=%s', (mid,)
    ).fetchone()
    if not r:
        abort(404)
    def _s(v): return str(v)[:10] if v else ''
    def _dt(v): return str(v)[:19] if v else ''
    from datetime import date as _date
    def _edad(fnac):
        if not fnac:
            return ''
        hoy = _date.today()
        try:
            d = fnac if hasattr(fnac, 'year') else _date.fromisoformat(str(fnac)[:10])
            return hoy.year - d.year - ((hoy.month, hoy.day) < (d.month, d.day))
        except Exception:
            return ''
    def _firmantes(json_str):
        if not json_str:
            return '', ''
        try:
            firmas = _json.loads(json_str)
        except Exception:
            return '', ''
        return (firmas.get('op') or {}).get('usuario', ''), \
               (firmas.get('coord') or firmas.get('admin') or {}).get('usuario', '')
    op_firma, coord_firma = _firmantes(r['firmas_aprobacion'])
    headers = ['Folio', 'Nombre', 'Fecha de nacimiento', 'Edad',
               'País de origen', 'Grupo de población',
               'Fecha de atención', 'Género', 'Departamento/Estado',
               'Estado civil', 'Teléfono', 'Registrado por',
               'Firmado por (operativo)', 'Firmado por (coordinador/admin)',
               'Fecha de creación']
    data = [(
        r['folio'], r['nombre'],
        _s(r['fecha_nacimiento']), _edad(r['fecha_nacimiento']),
        r['pais_origen'], r['grupo_poblacion'],
        _s(r['fecha_atencion']), r['genero'], r['departamento_estado'],
        r['estado_civil'], r['telefono_contacto'], r['registrado_por'],
        op_firma, coord_firma,
        _dt(r['creado']),
    )]
    buf = _make_excel(headers, data)
    folio_safe = (r['folio'] or 'migrante').replace('/', '-')
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'{folio_safe}.xlsx')


@app.route('/admin/descargar/eliminados')
@verificar_certificado
def descargar_excel_eliminados():
    if g.rol != 'admin':
        abort(403)
    fecha_inicio = request.args.get('fecha_inicio', '').strip() or None
    fecha_fin    = request.args.get('fecha_fin',    '').strip() or None

    sql    = 'SELECT * FROM historial_eliminaciones WHERE 1=1'
    params = []
    if fecha_inicio:
        sql += ' AND fecha_eliminacion >= %s'; params.append(fecha_inicio)
    if fecha_fin:
        sql += ' AND fecha_eliminacion <= %s'; params.append(fecha_fin + ' 23:59:59')
    sql += ' ORDER BY id DESC'

    db   = obtener_db()
    rows = _exec(db, sql, params).fetchall()

    def _sv(v): return str(v)[:19] if v else ''

    headers = ['Fecha eliminación', 'Folio', 'Nombre', 'Tipo',
               'Solicitado por', 'Motivo', 'Firma coordinador', 'Aprobado por']
    data = [
        (_sv(r['fecha_eliminacion']), r['folio'] or '', r['nombre_migrante'] or '',
         r['tipo'] or '', r['solicitado_por'] or '', r['motivo'] or '',
         'Sí' if r['firma_coord'] else 'No', r['aprobado_por'] or '')
        for r in rows
    ]
    buf = _make_excel(headers, data)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='eliminados.xlsx')


@app.route('/admin/descargar/log')
@verificar_certificado
def descargar_excel_log():
    if g.rol != 'admin':
        abort(403)
    fecha_inicio = request.args.get('fecha_inicio', '').strip() or None
    fecha_fin    = request.args.get('fecha_fin',    '').strip() or None

    sql    = ('SELECT fecha, usuario, rol, ip_origen, resultado, detalle '
              'FROM log_certificados WHERE 1=1')
    params = []
    if fecha_inicio:
        sql += ' AND fecha >= %s'; params.append(fecha_inicio)
    if fecha_fin:
        sql += ' AND fecha <= %s'; params.append(fecha_fin + ' 23:59:59')
    sql += ' ORDER BY id DESC'

    db   = obtener_db()
    rows = _exec(db, sql, params).fetchall()

    headers = ['Fecha', 'Usuario', 'Rol', 'IP origen', 'Resultado', 'Detalle']
    data = [
        (str(r['fecha'])[:19] if r['fecha'] else '',
         r['usuario'], r['rol'], r['ip_origen'], r['resultado'], r['detalle'])
        for r in rows
    ]
    buf = _make_excel(headers, data)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='log_auditoria.xlsx')


# ── GESTIÓN DE USUARIOS ───────────────────────────────────────

@app.route('/admin/api/usuarios')
@verificar_certificado
def admin_api_usuarios():
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    rows = _exec(db,
        'SELECT usuario, nombre, rol, activo, debe_cambiar_pwd, creado,'
        ' telefono, curp, fecha_nacimiento, genero, area, observaciones,'
        ' COALESCE(bloqueado, 0) AS bloqueado,'
        ' COALESCE(intentos_fallidos, 0) AS intentos_fallidos'
        ' FROM usuarios ORDER BY rol, nombre'
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('creado') and hasattr(d['creado'], 'isoformat'):
            d['creado'] = d['creado'].isoformat()
        if d.get('fecha_nacimiento') and hasattr(d['fecha_nacimiento'], 'isoformat'):
            d['fecha_nacimiento'] = d['fecha_nacimiento'].isoformat()[:10]
        result.append(d)
    return jsonify(result)


@app.route('/admin/api/usuarios/reset_pwd', methods=['POST'])
@verificar_certificado
def admin_reset_pwd():
    if g.rol != 'admin':
        abort(403)
    usuario = request.form.get('usuario', '').strip()
    if not usuario:
        return jsonify(error='Usuario requerido'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT nombre FROM usuarios WHERE usuario=%s', (usuario,)).fetchone()
    if not user:
        return jsonify(error='Usuario no encontrado'), 404
    pwd_temp = generar_contrasena_temporal()
    _exec(db,
        'UPDATE usuarios SET password_hash=%s, debe_cambiar_pwd=1 WHERE usuario=%s',
        (generate_password_hash(pwd_temp), usuario)
    )
    db.commit()
    log_evento('pwd_cambiada', usuario=usuario, rol=g.rol,
               detalle=f'Reset forzado por {g.usuario}')
    return jsonify(ok=True, pwd_temporal=pwd_temp)


@app.route('/admin/api/usuarios/toggle_activo', methods=['POST'])
@verificar_certificado
def admin_toggle_activo():
    if g.rol != 'admin':
        abort(403)
    usuario = request.form.get('usuario', '').strip()
    if not usuario:
        return jsonify(error='Usuario requerido'), 400
    if usuario == g.usuario:
        return jsonify(error='No puedes desactivar tu propia cuenta'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT activo FROM usuarios WHERE usuario=%s', (usuario,)).fetchone()
    if not user:
        return jsonify(error='Usuario no encontrado'), 404
    nuevo = 0 if user['activo'] else 1
    _exec(db, 'UPDATE usuarios SET activo=%s WHERE usuario=%s', (nuevo, usuario))
    db.commit()
    accion = 'activada' if nuevo else 'desactivada'
    log_evento('rol_modificado', usuario=usuario, rol=g.rol,
               detalle=f'Cuenta {accion} por {g.usuario}')
    return jsonify(ok=True, activo=nuevo)


@app.route('/admin/api/usuarios/eliminar', methods=['POST'])
@verificar_certificado
def admin_eliminar_usuario():
    if g.rol != 'admin':
        abort(403)
    usuario = request.form.get('usuario', '').strip()
    if not usuario:
        return jsonify(error='Usuario requerido'), 400
    if usuario == g.usuario:
        return jsonify(error='No puedes eliminar tu propia cuenta'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT activo, nombre, rol FROM usuarios WHERE usuario=%s',
                 (usuario,)).fetchone()
    if not user:
        return jsonify(error='Usuario no encontrado'), 404
    if user['activo']:
        return jsonify(error='Solo se pueden eliminar cuentas desactivadas'), 400
    try:
        _exec(db, 'DELETE FROM solicitudes_pwd WHERE usuario=%s', (usuario,))
        _exec(db, 'DELETE FROM usuarios WHERE usuario=%s', (usuario,))
        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify(error=f'Error al eliminar usuario: {str(e)}'), 500
    log_evento('rol_modificado', usuario=usuario, rol=user['rol'],
               detalle=f'Usuario eliminado permanentemente por {g.usuario}')
    return jsonify(ok=True)


@app.route('/admin/api/usuarios/desbloquear', methods=['POST'])
@verificar_certificado
def admin_desbloquear_usuario():
    if g.rol != 'admin':
        abort(403)
    usuario = request.form.get('usuario', '').strip()
    if not usuario:
        return jsonify(error='Usuario requerido'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT nombre FROM usuarios WHERE usuario=%s', (usuario,)).fetchone()
    if not user:
        return jsonify(error='Usuario no encontrado'), 404
    try:
        _exec(db, 'UPDATE usuarios SET bloqueado=0, intentos_fallidos=0 WHERE usuario=%s', (usuario,))
        db.commit()
    except Exception:
        _exec(db, 'UPDATE usuarios SET activo=1 WHERE usuario=%s', (usuario,))
        db.commit()
    log_evento('acceso_denegado', usuario=usuario, rol=g.rol,
               detalle=f'Cuenta desbloqueada por {g.usuario}')
    return jsonify(ok=True)


@app.route('/admin/api/usuarios/editar', methods=['POST'])
@verificar_certificado
def admin_editar_usuario():
    if g.rol != 'admin':
        abort(403)
    usuario     = request.form.get('usuario', '').strip()
    nombre      = request.form.get('nombre', '').strip()
    rol         = request.form.get('rol', '').strip()
    telefono    = request.form.get('telefono', '').strip() or None
    curp        = request.form.get('curp', '').strip() or None
    fecha_nac   = request.form.get('fecha_nacimiento', '').strip() or None
    genero      = request.form.get('genero', '').strip() or None
    area        = request.form.get('area', '').strip() or None
    observ      = request.form.get('observaciones', '').strip() or None
    if not usuario or not nombre or not rol:
        return jsonify(error='Datos incompletos'), 400
    if rol not in ('admin', 'coord', 'op', 'voluntario'):
        return jsonify(error='Rol no válido'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT id, rol, serial_cert FROM usuarios WHERE usuario=%s',
                 (usuario,)).fetchone()
    if not user:
        return jsonify(error='Usuario no encontrado'), 404

    rol_anterior    = user['rol']
    serial_anterior = user['serial_cert']

    # ── Solo datos personales / mismo rol ────────────────────
    if rol == rol_anterior:
        _exec(db,
            'UPDATE usuarios SET nombre=%s, telefono=%s, curp=%s,'
            ' fecha_nacimiento=%s, genero=%s, area=%s, observaciones=%s'
            ' WHERE usuario=%s',
            (nombre, telefono, curp, fecha_nac, genero, area, observ, usuario)
        )
        db.commit()
        log_evento('rol_modificado', usuario=usuario, rol=g.rol,
                   detalle=f'Perfil editado por {g.usuario}: nombre="{nombre}"')
        return jsonify(ok=True)

    # ── Revocar certificado anterior si venía de coord/admin ──
    def _revocar_serial_anterior(serial):
        if not serial:
            return
        cert_row = _exec(db, 'SELECT id FROM certificados WHERE serial=%s AND estado=%s',
                         (serial, 'vigente')).fetchone()
        if cert_row:
            _exec(db, "UPDATE certificados SET estado='revocado' WHERE serial=%s", (serial,))
            _exec(db,
                'INSERT IGNORE INTO certificados_revocados'
                ' (serial, usuario, rol, revocado_por, motivo) VALUES (%s,%s,%s,%s,%s)',
                (serial, usuario, rol_anterior, g.usuario, f'Cambio de rol a {rol}')
            )
            log_evento('cert_revocado_admin', usuario=usuario, rol=rol_anterior,
                       serial=serial,
                       detalle=f'Revocado automáticamente al cambiar rol a {rol}')

    # ── Cambio a coordinador: generar P12 ────────────────────
    if rol == 'coord':
        pwd_temporal = generar_contrasena_temporal()
        pwd_p12      = generar_contrasena_temporal()
        try:
            p12_bytes, serial, cert = generar_p12_real(usuario, nombre, 'coord', pwd_p12)
        except Exception as e:
            return jsonify(error=f'Error al generar certificado: {str(e)}'), 500

        crt_bytes = cert.public_bytes(serialization.Encoding.PEM)
        nombre_p12 = f"coord_{nombre.lower().replace(' ', '_')}"
        readme_texto = (
            f'Instrucciones para instalar certificado de Casa Monarca\n\n'
            f'1. Ingresa con tu email y contraseña temporal en la pantalla de login.\n\n'
            f'2. Selecciona el archivo {nombre_p12}.p12 con esta contraseña: {pwd_p12}\n\n'
            f'⚠️ No copies estos archivos fuera del USB.\n'
            f'⚠️ La contraseña es personal y solo se muestra aquí.\n'
        )

        _revocar_serial_anterior(serial_anterior)

        fecha     = datetime.now().strftime('%Y-%m-%d')
        fecha_exp = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        _exec(db,
            'INSERT INTO certificados'
            ' (serial, usuario, nombre, rol, fecha_emision, fecha_expiracion, estado, emitido_por)'
            ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
            (serial, usuario, nombre, 'coord', fecha, fecha_exp, 'vigente', g.usuario)
        )
        _exec(db,
            'UPDATE usuarios SET nombre=%s, rol=%s, password_hash=%s,'
            ' debe_cambiar_pwd=1, serial_cert=%s WHERE usuario=%s',
            (nombre, 'coord', generate_password_hash(pwd_temporal), serial, usuario)
        )
        db.commit()

        nombre_seguro = nombre_p12
        session['pwd_login']        = pwd_temporal
        session['p12_b64']          = base64.b64encode(p12_bytes).decode('utf-8')
        session['crt_b64']          = base64.b64encode(crt_bytes).decode('utf-8')
        session['readme_txt']       = readme_texto
        session['p12_nombre']       = nombre_seguro
        session['p12_display']      = nombre
        session['p12_email']        = usuario
        session['p12_rol']          = 'coord'
        session['p12_serial']       = serial
        session['p12_pwd_mostrada'] = False

        log_evento('cert_emitido', usuario=usuario, rol='coord', serial=serial,
                   detalle=f'Emitido por cambio de rol desde {g.usuario}')
        log_evento('rol_modificado', usuario=usuario, rol=g.rol,
                   detalle=f'Rol cambiado a coord por {g.usuario}; nuevo cert {serial}')
        return jsonify(ok=True, accion='certificado',
                       redirect_url=url_for('entregar_certificado'))

    # ── Cambio a op o voluntario ──────────────────────────────
    pwd_temporal = generar_contrasena_temporal()
    _revocar_serial_anterior(serial_anterior)
    _exec(db,
        'UPDATE usuarios SET nombre=%s, rol=%s, password_hash=%s,'
        ' debe_cambiar_pwd=1, serial_cert=NULL WHERE usuario=%s',
        (nombre, rol, generate_password_hash(pwd_temporal), usuario)
    )
    db.commit()
    log_evento('rol_modificado', usuario=usuario, rol=g.rol,
               detalle=f'Rol cambiado a {rol} por {g.usuario}; contraseña temporal asignada')
    return jsonify(ok=True, accion='pwd_temporal', pwd_temporal=pwd_temporal)


@app.route('/solicitar_reset_pwd', methods=['POST'])
def solicitar_reset_pwd():
    usuario = request.form.get('usuario', '').strip().lower()
    if not usuario:
        return jsonify(error='Ingresa tu email de acceso'), 400
    db = obtener_db()
    user = _exec(db, 'SELECT id FROM usuarios WHERE usuario=%s', (usuario,)).fetchone()
    if user:
        existe = _exec(db,
            "SELECT id FROM solicitudes_pwd WHERE usuario=%s AND estado='pendiente'",
            (usuario,)
        ).fetchone()
        if not existe:
            _exec(db, 'INSERT INTO solicitudes_pwd (usuario) VALUES (%s)', (usuario,))
            db.commit()
            log_evento('pwd_reset_solicitado', usuario=usuario,
                       detalle='Solicitud desde pantalla de login')
    return jsonify(ok=True)


@app.route('/admin/api/solicitudes_pwd')
@verificar_certificado
def admin_solicitudes_pwd():
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    rows = _exec(db,
        "SELECT id, usuario, estado, solicitado_en, resuelto_por, resuelto_en"
        " FROM solicitudes_pwd ORDER BY solicitado_en DESC LIMIT 100"
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in ('solicitado_en', 'resuelto_en'):
            if d.get(k) and hasattr(d[k], 'isoformat'):
                d[k] = d[k].isoformat()
        result.append(d)
    return jsonify(result)


@app.route('/admin/api/solicitudes_pwd/<int:sid>/resolver', methods=['POST'])
@verificar_certificado
def admin_resolver_reset_pwd(sid):
    if g.rol != 'admin':
        abort(403)
    aprobar = request.form.get('accion') == 'aprobar'
    db = obtener_db()
    sol = _exec(db,
        "SELECT usuario, estado FROM solicitudes_pwd WHERE id=%s", (sid,)
    ).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    if sol['estado'] != 'pendiente':
        return jsonify(error='Esta solicitud ya fue resuelta'), 400

    pwd_temp = None
    if aprobar:
        pwd_temp = generar_contrasena_temporal()
        _exec(db,
            'UPDATE usuarios SET password_hash=%s, debe_cambiar_pwd=1 WHERE usuario=%s',
            (generate_password_hash(pwd_temp), sol['usuario'])
        )
        resultado_log = 'pwd_reset_aprobado'
    else:
        resultado_log = 'pwd_reset_rechazado'

    _exec(db,
        'UPDATE solicitudes_pwd SET estado=%s, resuelto_por=%s, resuelto_en=NOW() WHERE id=%s',
        ('aprobada' if aprobar else 'rechazada', g.usuario, sid)
    )
    db.commit()
    log_evento(resultado_log, usuario=sol['usuario'], rol=g.rol,
               detalle=f'{"Aprobado" if aprobar else "Rechazado"} por {g.usuario}')

    resp = dict(ok=True)
    if pwd_temp:
        resp['pwd_temporal'] = pwd_temp
    return jsonify(**resp)


# ── EMITIR NUEVO CERTIFICADO ──────────────────────────────────
@app.route('/admin/nuevo_voluntario', methods=['POST'])
@verificar_certificado
def nuevo_voluntario():
    if g.rol != 'admin':
        abort(403)

    try:
        nombre = sanitizar_nombre(request.form.get('nombre', ''))
        email  = request.form.get('correo', '').strip().lower()
        rol    = request.form.get('rol', '')

        if not nombre:
            return jsonify(error='El nombre no puede estar vacío'), 400
        if not email:
            return jsonify(error='El email es obligatorio'), 400
        if rol not in ('op', 'coord', 'admin', 'voluntario'):
            return jsonify(error='Rol inválido'), 400

        correo   = request.form.get('correo', '').strip() or None
        telefono = request.form.get('telefono', '').strip() or None
        curp     = request.form.get('curp', '').strip() or None
        fnac     = request.form.get('fecha_nac', '').strip() or None
        genero   = request.form.get('genero', '').strip() or None
        area     = request.form.get('area', '').strip() or None
        observ   = request.form.get('observaciones', '').strip() or None

        db = obtener_db()
        existe = _exec(db, 'SELECT id FROM usuarios WHERE usuario=%s', (email,)).fetchone()
        if existe:
            return jsonify(error=f'El email {email} ya está registrado'), 400

        pwd_temporal = generar_contrasena_temporal()

        # Voluntario y operador: solo cuenta con contraseña temporal, sin certificado
        if rol in ('voluntario', 'op'):
            _exec(db,
                'INSERT INTO usuarios (usuario, nombre, rol, password_hash,'
                ' debe_cambiar_pwd, correo, telefono, curp, fecha_nacimiento,'
                ' genero, area, observaciones)'
                ' VALUES (%s,%s,%s,%s,1,%s,%s,%s,%s,%s,%s,%s)',
                (email, nombre, rol, generate_password_hash(pwd_temporal),
                 correo, telefono, curp, fnac, genero, area, observ)
            )
            db.commit()
            session['pwd_login']        = pwd_temporal
            session['p12_nombre']       = f"{rol}_{nombre.lower().replace(' ', '_')}"
            session['p12_display']      = nombre    
            session['p12_email']        = email
            session['p12_rol']          = rol
            session['p12_serial']       = ''
            session['p12_pwd_mostrada'] = False
            log_evento('rol_modificado', usuario=email, rol=rol,
                       detalle=f'Alta voluntario por {g.usuario}')
            return jsonify(ok=True, nombre=nombre, email=email, rol=rol,
                           redirect_url=url_for('entregar_certificado'))

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error=f'Error al crear usuario: {str(e)}'), 500

    # Roles con certificado (admin, coord únicamente)
    pwd_p12 = pwd_temporal
    if rol in ('admin', 'coord'):
        pwd_p12 = generar_contrasena_temporal()

    try:
        p12_bytes, serial, cert = generar_p12_real(email, nombre, rol, pwd_p12)
    except Exception as e:
        return jsonify(error=f'Error al generar el certificado .p12: {str(e)}'), 500

    # Generar bytes del .crt en memoria (sin escribir a disco)
    crt_bytes = cert.public_bytes(serialization.Encoding.PEM)

    # Generar README en memoria
    nombre_p12 = f"{rol}_{nombre.lower().replace(' ', '_')}"
    readme_texto = f"""Instrucciones para instalar certificado de Casa Monarca

1. Ingresa con tu email y contraseña (contraseña temporal si es la primera vez)
   en la pantalla de login de Casa Monarca.

2. Selecciona el archivo {nombre_p12}.p12 en tu navegador con esta contraseña: {pwd_p12}

⚠️ No copies estos archivos fuera del USB.
⚠️ La contraseña es personal y solo se muestra aquí.
"""

    fecha = datetime.now().strftime('%Y-%m-%d')
    fecha_exp = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')

    _exec(db,
        'INSERT INTO certificados'
        ' (serial, usuario, nombre, rol, fecha_emision, fecha_expiracion, estado, emitido_por)'
        ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
        (serial, email, nombre, rol, fecha, fecha_exp, 'vigente', g.usuario)
    )

    existente = _exec(db, 'SELECT id FROM usuarios WHERE usuario=%s', (email,)).fetchone()
    if existente is None:
        _exec(db,
            'INSERT INTO usuarios (usuario, nombre, rol, password_hash,'
            ' debe_cambiar_pwd, correo, telefono, curp, fecha_nacimiento,'
            ' genero, area, observaciones, serial_cert)'
            ' VALUES (%s,%s,%s,%s,1,%s,%s,%s,%s,%s,%s,%s,%s)',
            (email, nombre, rol,
             generate_password_hash(pwd_temporal),
             email, telefono, curp, fnac, genero, area, observ, serial)
        )
    else:
        _exec(db,
            'UPDATE usuarios SET nombre=%s, rol=%s, password_hash=%s,'
            ' debe_cambiar_pwd=1, serial_cert=%s WHERE usuario=%s',
            (nombre, rol, generate_password_hash(pwd_temporal),
             serial, email)
        )
    db.commit()

    log_evento('cert_emitido', usuario=email, rol=rol, serial=serial,
               detalle=f'Emitido por {g.usuario} para {nombre}')

    # Guardar en sesión el nombre correcto del .p12
    nombre_seguro = f"{rol}_{nombre.lower().replace(' ', '_')}"
    session['pwd_login'] = pwd_temporal
    session['p12_b64'] = base64.b64encode(p12_bytes).decode('utf-8')
    session['crt_b64'] = base64.b64encode(crt_bytes).decode('utf-8')
    session['readme_txt'] = readme_texto
    session['p12_nombre'] = nombre_seguro
    session['p12_display'] = nombre    
    session['p12_email'] = email
    session['p12_rol'] = rol
    session['p12_serial'] = serial
    session['p12_pwd_mostrada'] = False

    return jsonify(
        ok=True,
        nombre=nombre,
        email=email,
        rol=rol,
        serial=serial,
        redirect_url=url_for('entregar_certificado')
    )

# ── PANTALLA DE ENTREGA ───────────────────────────────────────
@app.route('/admin/entregar_certificado')
@verificar_certificado
def entregar_certificado():
    print(">>> Renderizando entregar_certificado")
    if g.rol != 'admin':
        abort(403)

    nombre = session.get('p12_display')
    rol = session.get('p12_rol')
    serial = session.get('p12_serial')

    if not nombre:
        return redirect(url_for('panel_admin'))

    pwd_login = None
    if not session.get('p12_pwd_mostrada', False):
        pwd_login = session.get('pwd_login')
        session['p12_pwd_mostrada'] = True

    return render_template(
        'entregar_cert.html',
        nombre=nombre,
        rol=rol,
        rol_legible=badge_rol(rol),
        serial=serial,
        pwd_login=pwd_login,
        correo=session.get('p12_email', '')
    )

# ── DESCARGA REAL DEL .P12 ────────────────────────────────────
@app.route('/admin/descargar_p12')
@verificar_certificado
def descargar_p12():
    if g.rol != 'admin':
        abort(403)

    p12_b64      = session.get('p12_b64')
    crt_b64      = session.get('crt_b64')
    readme_txt   = session.get('readme_txt', '')
    nombre_seguro = session.get('p12_nombre', 'certificado')
    email        = session.get('p12_email', 'usuario@casamonarca.org')

    if not p12_b64:
        abort(404, description='No hay certificado disponible para descargar.')

    try:
        p12_bytes = base64.b64decode(p12_b64)
    except Exception:
        abort(500, description='No se pudo reconstruir el archivo .p12.')

    # Construir el ZIP completamente en memoria
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{nombre_seguro}.p12", p12_bytes)
        if crt_b64:
            zf.writestr(f"{email}.crt", base64.b64decode(crt_b64))
        if readme_txt:
            zf.writestr(f"README_{email}.txt", readme_txt.encode('utf-8'))
    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f"{nombre_seguro}_usb.zip"
    )

# ── REVOCAR CERTIFICADO ───────────────────────────────────────
@app.route('/admin/revocar/<int:cert_id>', methods=['POST'])
@verificar_certificado
def revocar_certificado(cert_id):
    if g.rol != 'admin':
        abort(403)

    db = obtener_db()
    cert = _exec(db,
        'SELECT * FROM certificados WHERE id=%s', (cert_id,)
    ).fetchone()
    if cert is None:
        return jsonify(error='Certificado no encontrado'), 404
    if cert['estado'] == 'revocado':
        return jsonify(error='Certificado ya estaba revocado'), 400

    motivo = request.form.get('motivo', 'Revocado desde panel admin')
    _exec(db, "UPDATE certificados SET estado='revocado' WHERE id=%s", (cert_id,))
    _exec(db,
        'INSERT IGNORE INTO certificados_revocados'
        ' (serial, usuario, rol, revocado_por, motivo) VALUES (%s,%s,%s,%s,%s)',
        (cert['serial'], cert['usuario'], cert['rol'], g.usuario, motivo)
    )
    _exec(db, 'UPDATE usuarios SET activo=0 WHERE usuario=%s', (cert['usuario'],))
    db.commit()

    log_evento('cert_revocado_admin', usuario=cert['usuario'], rol=cert['rol'],
               serial=cert['serial'], detalle=motivo)
    return jsonify(ok=True)

# ── ELIMINAR CERTIFICADO ──────────────────────────────────────
@app.route('/admin/certificados/<int:cert_id>/eliminar', methods=['POST'])
@verificar_certificado
def eliminar_certificado(cert_id):
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    cert = _exec(db, 'SELECT * FROM certificados WHERE id=%s', (cert_id,)).fetchone()
    if cert is None:
        return jsonify(error='Certificado no encontrado'), 404
    try:
        if cert['estado'] == 'vigente':
            _exec(db, 'UPDATE usuarios SET activo=0 WHERE usuario=%s', (cert['usuario'],))
            _exec(db,
                'INSERT IGNORE INTO certificados_revocados'
                ' (serial, usuario, rol, revocado_por, motivo) VALUES (%s,%s,%s,%s,%s)',
                (cert['serial'], cert['usuario'], cert['rol'], g.usuario, 'Eliminado desde panel admin')
            )
        _exec(db, 'DELETE FROM certificados WHERE id=%s', (cert_id,))
        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify(error=f'Error al eliminar: {e}'), 500
    log_evento('cert_revocado_admin', usuario=cert['usuario'], rol=cert['rol'],
               serial=cert['serial'], detalle=f'Certificado eliminado por {g.usuario}')
    return jsonify(ok=True)

# ── RENOVAR CERTIFICADO ───────────────────────────────────────
@app.route('/admin/certificados/<int:cert_id>/renovar', methods=['POST'])
@verificar_certificado
def renovar_certificado(cert_id):
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    cert = _exec(db, 'SELECT * FROM certificados WHERE id=%s', (cert_id,)).fetchone()
    if not cert:
        return jsonify(error='Certificado no encontrado'), 404
    if cert['estado'] == 'revocado':
        return jsonify(error='El certificado ya está revocado'), 400

    email  = cert['usuario']
    nombre = cert['nombre'] or cert['usuario']
    rol    = cert['rol']

    pwd_temporal = generar_contrasena_temporal()
    pwd_p12      = generar_contrasena_temporal()
    try:
        p12_bytes, serial, cert_obj = generar_p12_real(email, nombre, rol, pwd_p12)
    except Exception as e:
        return jsonify(error=f'Error al generar certificado: {e}'), 500

    crt_bytes = cert_obj.public_bytes(serialization.Encoding.PEM)
    nombre_p12 = f"{rol}_{nombre.lower().replace(' ', '_')}"
    readme_texto = (
        f'Instrucciones para instalar certificado de Casa Monarca\n\n'
        f'1. Ingresa con tu email y contraseña temporal en la pantalla de login.\n\n'
        f'2. Selecciona el archivo {nombre_p12}.p12 con esta contraseña: {pwd_p12}\n\n'
        f'⚠️ No copies estos archivos fuera del USB.\n'
        f'⚠️ La contraseña es personal y solo se muestra aquí.\n'
    )

    # Revocar cert anterior
    _exec(db, "UPDATE certificados SET estado='revocado' WHERE id=%s", (cert_id,))
    _exec(db,
        'INSERT IGNORE INTO certificados_revocados'
        ' (serial, usuario, rol, revocado_por, motivo) VALUES (%s,%s,%s,%s,%s)',
        (cert['serial'], email, rol, g.usuario, 'Renovación de certificado')
    )

    fecha     = datetime.now().strftime('%Y-%m-%d')
    fecha_exp = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
    _exec(db,
        'INSERT INTO certificados'
        ' (serial, usuario, nombre, rol, fecha_emision, fecha_expiracion, estado, emitido_por)'
        ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
        (serial, email, nombre, rol, fecha, fecha_exp, 'vigente', g.usuario)
    )
    _exec(db,
        'UPDATE usuarios SET password_hash=%s, debe_cambiar_pwd=1, activo=1, serial_cert=%s WHERE usuario=%s',
        (generate_password_hash(pwd_temporal), serial, email)
    )
    db.commit()

    log_evento('cert_emitido', usuario=email, rol=rol, serial=serial,
               detalle=f'Renovado por {g.usuario}')

    nombre_seguro = f"{rol}_{nombre.lower().replace(' ', '_')}"
    session['pwd_login'] = pwd_temporal
    session['p12_b64']   = base64.b64encode(p12_bytes).decode('utf-8')
    session['crt_b64']   = base64.b64encode(crt_bytes).decode('utf-8')
    session['readme_txt'] = readme_texto
    session['p12_nombre'] = nombre_seguro
    session['p12_display'] = nombre
    session['p12_email']   = email
    session['p12_rol']     = rol
    session['p12_serial']  = serial
    session['p12_pwd_mostrada'] = False

    return jsonify(ok=True, redirect_url=url_for('entregar_certificado'))


# ── PANEL DE COORDINADORES Y OPERADORES ───────────────────────
@app.route('/panel')
@verificar_certificado
def panel_usuario():
    if g.rol == 'admin':
        return redirect(url_for('panel_admin'))
    if g.rol == 'voluntario':
        return redirect(url_for('panel_voluntario'))
    return render_template('panel_usuario.html',
                           nombre=g.nombre, rol=g.rol, usuario=g.usuario,
                           rol_legible=badge_rol(g.rol))

# ── PANEL DE VOLUNTARIOS ──────────────────────────────────────
@app.route('/voluntario')
@verificar_certificado
def panel_voluntario():
    if g.rol != 'voluntario':
        abort(403)
    return render_template('voluntario.html',
                           nombre=g.nombre, rol=g.rol, usuario=g.usuario,
                           rol_legible='Voluntario')

@app.route('/panel/api/certificados')
@verificar_certificado
def panel_api_certificados():
    """Lista de certificados (solo lectura, sin botón de revocar).
    Coord y op pueden VER pero no modificar."""
    refrescar_expirados()
    db = obtener_db()
    rows = _exec(db,
        'SELECT id, serial, usuario, rol, fecha_emision, estado'
        ' FROM certificados ORDER BY id DESC'
    ).fetchall()
    lista = []
    for c in rows:
        lista.append({
            'id': c['id'],
            'nombre': c['usuario'],
            'rol': c['rol'],
            'serial': c['serial'],
            'fecha': c['fecha_emision'],
            'estado': c['estado'],
        })
    return jsonify(lista)

@app.route('/panel/api/mi_certificado')
@verificar_certificado
def panel_api_mi_certificado():
    """Devuelve los datos del cert más reciente del usuario logueado."""
    refrescar_expirados()
    db = obtener_db()
    # Buscar por nombre (como está guardado en la tabla certificados)
    cert = _exec(db,
        'SELECT serial, usuario, rol, fecha_emision, fecha_expiracion, estado'
        ' FROM certificados WHERE usuario=%s ORDER BY id DESC LIMIT 1',
        (g.nombre,)
    ).fetchone()
    if not cert:
        cert = _exec(db,
            'SELECT serial, usuario, rol, fecha_emision, fecha_expiracion, estado'
            ' FROM certificados WHERE usuario=%s ORDER BY id DESC LIMIT 1',
            (g.usuario,)
        ).fetchone()
    if not cert:
        return jsonify(error='Sin certificado asociado')
    return jsonify({
        'serial': cert['serial'],
        'nombre': cert['usuario'],
        'rol': cert['rol'],
        'fecha_emision': cert['fecha_emision'],
        'fecha_expiracion': cert['fecha_expiracion'] or '',
        'estado': cert['estado'],
    })

# ── PERFIL PROPIO DEL USUARIO ─────────────────────────────────
@app.route('/panel/api/mi_perfil')
@verificar_certificado
def panel_api_mi_perfil():
    db = obtener_db()
    u = _exec(db,
        'SELECT nombre, usuario, rol, telefono, curp, fecha_nacimiento,'
        '       genero, area FROM usuarios WHERE usuario=%s',
        (g.usuario,)
    ).fetchone()
    if not u:
        return jsonify(error='Usuario no encontrado'), 404
    d = dict(u)
    if d.get('fecha_nacimiento') and hasattr(d['fecha_nacimiento'], 'isoformat'):
        d['fecha_nacimiento'] = d['fecha_nacimiento'].isoformat()[:10]
    return jsonify(d)


# ── DIAGNÓSTICO DB ────────────────────────────────────────────
@app.route('/api/debug/columnas_migrantes')
@verificar_certificado
def debug_columnas():
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    rows = _exec(db, 'DESCRIBE migrantes').fetchall()
    return jsonify([dict(r) for r in rows])

def _validar_fecha_rango(fecha_str, nombre_campo):
    """None si es válida; mensaje de error si no lo es."""
    if not fecha_str:
        return None
    try:
        from datetime import date as _date
        d = _date.fromisoformat(str(fecha_str)[:10])
        hoy = _date.today()
        if d > hoy:
            return f'{nombre_campo} no puede ser una fecha futura'
        limite = _date(hoy.year - 100, hoy.month, hoy.day)
        if d < limite:
            return f'{nombre_campo} no puede ser hace más de 100 años'
    except ValueError:
        return f'{nombre_campo} tiene un formato inválido'
    return None


# ── REGISTROS DE MIGRANTES ────────────────────────────────────
@app.route('/api/migrantes')
@verificar_certificado
def api_migrantes_lista():
    db = obtener_db()
    rows = _exec(db,
        'SELECT id, folio, nombre, fecha_nacimiento, pais_origen,'
        ' fecha_atencion, genero, departamento_estado, estado_civil,'
        ' grupo_poblacion, telefono_contacto, registrado_por, creado, actualizado'
        ' FROM migrantes ORDER BY id DESC'
    ).fetchall()

    def _fstr(v):
        return v.isoformat()[:10] if hasattr(v, 'isoformat') else (str(v)[:10] if v else None)

    lista = []
    for r in rows:
        m = dict(r)
        m['fecha_nacimiento'] = _fstr(m.get('fecha_nacimiento'))
        m['fecha_atencion']   = _fstr(m.get('fecha_atencion'))
        lista.append(m)
    return jsonify(lista)


@app.route('/api/migrantes', methods=['POST'])
@verificar_certificado
def api_migrantes_crear():
    if g.rol not in ('admin', 'coord', 'op', 'voluntario'):
        abort(403)

    nombre_pila      = request.form.get('nombre_pila', '').strip()
    primer_apellido  = request.form.get('primer_apellido', '').strip()
    segundo_apellido = request.form.get('segundo_apellido', '').strip()
    fecha_atencion   = request.form.get('fecha_atencion', '').strip() or None
    telefono_contacto = request.form.get('telefono_contacto', '').strip() or None
    genero           = request.form.get('genero', '').strip() or None
    pais_origen      = request.form.get('pais_origen', '').strip()
    departamento_estado = request.form.get('departamento_estado', '').strip() or None
    estado_civil     = request.form.get('estado_civil', '').strip() or None
    fecha_nac        = request.form.get('fecha_nacimiento', '').strip() or None
    grupo_poblacion  = request.form.get('grupo_poblacion', '').strip() or None

    faltantes = [f for f, v in [
        ('nombre_pila', nombre_pila),
        ('primer_apellido', primer_apellido),
        ('pais_origen', pais_origen),
        ('fecha_nacimiento', fecha_nac),
        ('genero', genero),
        ('grupo_poblacion', grupo_poblacion),
    ] if not v]
    if faltantes:
        return jsonify(error=f'Campos obligatorios faltantes: {", ".join(faltantes)}'), 400

    err_fnac = _validar_fecha_rango(fecha_nac, 'Fecha de nacimiento')
    if err_fnac:
        return jsonify(error=err_fnac), 400
    err_fat = _validar_fecha_rango(fecha_atencion, 'Fecha de atención')
    if err_fat:
        return jsonify(error=err_fat), 400

    partes = [nombre_pila, primer_apellido]
    if segundo_apellido:
        partes.append(segundo_apellido)
    nombre = ' '.join(partes)

    db = obtener_db()
    import json as _json

    # ── Voluntario: crea ticket pendiente de validación por op ──
    if g.rol == 'voluntario':
        try:
            cur = _exec(db,
                'INSERT INTO solicitudes_registro_migrante'
                ' (nombre, fecha_nacimiento, pais_origen, fecha_atencion, genero,'
                '  departamento_estado, estado_civil, grupo_poblacion, telefono_contacto,'
                '  estado, enviado_por, rol_enviado_por)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (nombre, fecha_nac, pais_origen, fecha_atencion, genero,
                 departamento_estado, estado_civil, grupo_poblacion, telefono_contacto,
                 'pendiente_op', g.usuario, 'voluntario')
            )
        except Exception as e:
            return jsonify(error=f'Error al crear solicitud: {e}'), 500
        ticket_id = cur.lastrowid
        db.commit()
        log_evento('registro_mig_enviado_voluntario', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{ticket_id}: {nombre} · País: {pais_origen}')
        return jsonify(ok=True, pendiente=True, ticket_id=ticket_id)

    # ── Operativo: se auto-firma y crea ticket pendiente de coord ──
    if g.rol == 'op':
        ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            cur_tmp = _exec(db,
                'INSERT INTO solicitudes_registro_migrante'
                ' (nombre, fecha_nacimiento, pais_origen, fecha_atencion, genero,'
                '  departamento_estado, estado_civil, grupo_poblacion, telefono_contacto,'
                '  estado, enviado_por, rol_enviado_por)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (nombre, fecha_nac, pais_origen, fecha_atencion, genero,
                 departamento_estado, estado_civil, grupo_poblacion, telefono_contacto,
                 'pendiente_coord', g.usuario, 'op')
            )
        except Exception as e:
            return jsonify(error=f'Error al crear solicitud: {e}'), 500
        ticket_id = cur_tmp.lastrowid
        db.commit()
        priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
        fnac_str = str(fecha_nac) if fecha_nac else ''
        pais_str = pais_origen or ''
        mensaje_op = f'REGISTRO_OP|{ticket_id}|{nombre}|{fnac_str}|{pais_str}|{g.usuario}|{g.usuario}|{ahora}'
        firma_op = firmar_mensaje(priv_pem, mensaje_op)
        if not verificar_firma(pub_pem, mensaje_op, firma_op):
            db.rollback()
            return jsonify(error='Error al generar firma digital'), 500
        _exec(db,
            'UPDATE solicitudes_registro_migrante'
            ' SET op_validador=%s, firma_op=%s, op_pubkey=%s, mensaje_firmado_op=%s, op_validado_en=%s'
            ' WHERE id=%s',
            (g.usuario, firma_op, pub_pem, mensaje_op, ahora, ticket_id)
        )
        db.commit()
        log_evento('registro_mig_enviado_op_directo', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{ticket_id}: {nombre} · País: {pais_origen}')
        return jsonify(ok=True, pendiente=True, ticket_id=ticket_id)

    # ── Coord / Admin: inserción directa con firma propia ──
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        cur = _exec(db,
            'INSERT INTO migrantes'
            ' (folio, nombre, fecha_nacimiento, pais_origen,'
            '  fecha_atencion, genero, departamento_estado, estado_civil,'
            '  grupo_poblacion, telefono_contacto, registrado_por)'
            ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
            (
                'TEMP', nombre, fecha_nac, pais_origen,
                fecha_atencion, genero, departamento_estado, estado_civil,
                grupo_poblacion, telefono_contacto, g.usuario,
            )
        )
    except Exception as e:
        return jsonify(error=f'Error al guardar en base de datos: {e}'), 500

    nuevo_id = cur.lastrowid
    folio = f'MIG-{datetime.now().strftime("%Y%m")}-{nuevo_id:04d}'
    _exec(db, 'UPDATE migrantes SET folio=%s WHERE id=%s', (folio, nuevo_id))

    priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
    fnac_str = str(fecha_nac) if fecha_nac else ''
    pais_str = pais_origen or ''
    mensaje_directo = f'REGISTRO_DIRECTO|{nuevo_id}|{folio}|{nombre}|{pais_str}|{g.usuario}|{ahora}'
    firma_directa = firmar_mensaje(priv_pem, mensaje_directo)
    firmas_json = _json.dumps({
        g.rol: {
            'usuario': g.usuario,
            'mensaje': mensaje_directo,
            'firma': firma_directa,
            'pubkey': pub_pem,
        }
    })
    _exec(db, 'UPDATE migrantes SET firmas_aprobacion=%s WHERE id=%s', (firmas_json, nuevo_id))
    db.commit()

    log_evento('migrante_registrado', usuario=g.usuario, rol=g.rol,
               detalle=f'Nuevo migrante: {nombre} · País: {pais_origen} · Folio: {folio}')
    log_evento('registro_mig_insertado', usuario=g.usuario, rol=g.rol,
               detalle=f'Folio: {folio} · Registro directo por {g.rol}')
    return jsonify(ok=True, id=nuevo_id, folio=folio)


# ── WORKFLOW DE APROBACIÓN DE REGISTRO DE MIGRANTES ───────────

@app.route('/api/registro_migrante/pendientes_op', methods=['GET'])
@verificar_certificado
def api_registro_pendientes_op():
    if g.rol != 'op':
        return jsonify(error=f'Rol incorrecto: {g.rol}'), 403
    try:
        db = obtener_db()
        cur = _exec(db,
            "SELECT id, nombre, fecha_nacimiento, pais_origen, fecha_atencion,"
            "       genero, departamento_estado, estado_civil, grupo_poblacion,"
            "       telefono_contacto, enviado_por, rol_enviado_por, creado"
            " FROM solicitudes_registro_migrante"
            " WHERE estado='pendiente_op'"
            " ORDER BY creado ASC"
        )
        filas = cur.fetchall()
    except Exception as e:
        return jsonify(error=str(e)), 500
    resultado = []
    for f in filas:
        r = dict(f)
        for k in ('fecha_nacimiento', 'fecha_atencion', 'creado'):
            if r.get(k) and hasattr(r[k], 'isoformat'):
                r[k] = r[k].isoformat()
        resultado.append(r)
    return jsonify(resultado)


@app.route('/api/registro_migrante/historial_op', methods=['GET'])
@verificar_certificado
def api_registro_historial_op():
    if g.rol != 'op':
        return jsonify(error=f'Rol incorrecto: {g.rol}'), 403
    try:
        limite = int(request.args.get('limite', '10'))
    except ValueError:
        limite = 10
    try:
        db = obtener_db()
        sql = (
            "SELECT id, nombre, pais_origen, fecha_atencion, estado, op_validado_en"
            " FROM solicitudes_registro_migrante"
            " WHERE op_validador=%s AND estado IN ('pendiente_coord','aprobada','rechazada')"
            " ORDER BY op_validado_en DESC"
        )
        params = [g.usuario]
        if limite > 0:
            sql += " LIMIT %s"
            params.append(limite)
        filas = _exec(db, sql, params).fetchall()
    except Exception as e:
        return jsonify(error=str(e)), 500
    resultado = []
    for f in filas:
        r = dict(f)
        for k in ('fecha_atencion', 'op_validado_en'):
            if r.get(k) and hasattr(r[k], 'isoformat'):
                r[k] = r[k].isoformat()
        resultado.append(r)
    return jsonify(resultado)


@app.route('/api/registro_migrante/pendientes_coord', methods=['GET'])
@verificar_certificado
def api_registro_pendientes_coord():
    if g.rol != 'coord':
        abort(403)
    db = obtener_db()
    cur = _exec(db,
        "SELECT id, nombre, fecha_nacimiento, pais_origen, fecha_atencion,"
        "       genero, departamento_estado, estado_civil, grupo_poblacion,"
        "       telefono_contacto, enviado_por, rol_enviado_por,"
        "       op_validador, firma_op, op_validado_en, creado"
        " FROM solicitudes_registro_migrante"
        " WHERE estado='pendiente_coord'"
        " ORDER BY creado ASC"
    )
    filas = cur.fetchall()
    resultado = []
    for f in filas:
        r = dict(f)
        for k in ('fecha_nacimiento', 'fecha_atencion', 'op_validado_en', 'creado'):
            if r.get(k) and hasattr(r[k], 'isoformat'):
                r[k] = r[k].isoformat()
        r['tiene_firma_op'] = bool(r.get('firma_op'))
        r.pop('firma_op', None)
        resultado.append(r)
    return jsonify(resultado)


@app.route('/api/registro_migrante/historial', methods=['GET'])
@verificar_certificado
def api_registro_historial():
    if g.rol not in ('op', 'coord'):
        return jsonify(error=f'Rol incorrecto: {g.rol}'), 403
    try:
        db = obtener_db()
        if g.rol == 'op':
            cur = _exec(db,
                "SELECT id, nombre, pais_origen, fecha_atencion, estado,"
                "       creado, op_validado_en, coord_aprobado_en,"
                "       rechazado_por, motivo_rechazo, rechazado_en"
                " FROM solicitudes_registro_migrante"
                " WHERE (op_validador=%s AND estado IN ('pendiente_coord','aprobada','rechazada'))"
                "    OR (rechazado_por=%s AND op_validador IS NULL AND estado='rechazada')"
                " ORDER BY COALESCE(op_validado_en, rechazado_en) DESC LIMIT 100",
                (g.usuario, g.usuario)
            )
        else:
            cur = _exec(db,
                "SELECT id, nombre, pais_origen, fecha_atencion, estado,"
                "       creado, op_validado_en, coord_aprobado_en,"
                "       rechazado_por, motivo_rechazo, rechazado_en"
                " FROM solicitudes_registro_migrante"
                " WHERE (coord_aprobador=%s AND estado='aprobada')"
                "    OR (rechazado_por=%s AND op_validador IS NOT NULL AND estado='rechazada')"
                " ORDER BY COALESCE(coord_aprobado_en, rechazado_en) DESC LIMIT 100",
                (g.usuario, g.usuario)
            )
        filas = cur.fetchall()
    except Exception as e:
        return jsonify(error=str(e)), 500
    resultado = []
    for f in filas:
        r = dict(f)
        for k in ('creado', 'op_validado_en', 'coord_aprobado_en', 'rechazado_en', 'fecha_atencion'):
            if r.get(k) and hasattr(r[k], 'isoformat'):
                r[k] = r[k].isoformat()
        resultado.append(r)
    return jsonify(resultado)


@app.route('/api/registro_migrante/<int:sid>/detalle', methods=['GET'])
@verificar_certificado
def api_registro_detalle(sid):
    if g.rol not in ('op', 'coord', 'admin'):
        return jsonify(error=f'Rol incorrecto: {g.rol}'), 403
    try:
        db = obtener_db()
        cur = _exec(db,
            "SELECT id, nombre, fecha_nacimiento, pais_origen, fecha_atencion,"
            "       genero, departamento_estado, estado_civil, grupo_poblacion, telefono_contacto,"
            "       estado, enviado_por, rol_enviado_por, creado,"
            "       op_validador, op_validado_en, firma_op,"
            "       coord_aprobador, coord_aprobado_en, firma_coord,"
            "       rechazado_por, rechazado_en, motivo_rechazo"
            " FROM solicitudes_registro_migrante WHERE id=%s",
            (sid,)
        )
        ticket = cur.fetchone()
    except Exception as e:
        return jsonify(error=str(e)), 500
    if not ticket:
        return jsonify(error='Registro no encontrado'), 404
    r = dict(ticket)
    # Solo puede verlo quien participó en él (como validador, aprobador o quien rechazó)
    # El admin puede ver cualquier registro
    if g.rol == 'op':
        if r.get('op_validador') != g.usuario and r.get('rechazado_por') != g.usuario:
            return jsonify(error='Sin acceso a este registro'), 403
    if g.rol == 'coord':
        if r.get('coord_aprobador') != g.usuario and r.get('rechazado_por') != g.usuario:
            return jsonify(error='Sin acceso a este registro'), 403
    for k in ('fecha_nacimiento', 'fecha_atencion', 'creado',
              'op_validado_en', 'coord_aprobado_en', 'rechazado_en'):
        if r.get(k) and hasattr(r[k], 'isoformat'):
            r[k] = r[k].isoformat()
    # Truncar firmas para mostrar solo los primeros 48 caracteres
    if r.get('firma_op'):
        r['firma_op'] = r['firma_op'][:48] + '…'
    if r.get('firma_coord'):
        r['firma_coord'] = r['firma_coord'][:48] + '…'
    return jsonify(r)


@app.route('/admin/api/actividad_detalle/eliminacion/<int:sid>')
@verificar_certificado
def admin_actividad_detalle_eliminacion(sid):
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    sol = _exec(db,
        'SELECT s.*, m.folio, m.nombre AS nombre_migrante'
        ' FROM solicitudes_eliminacion s'
        ' LEFT JOIN migrantes m ON m.id = s.migrante_id'
        ' WHERE s.id=%s', (sid,)
    ).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    r = dict(sol)
    # Si el migrante ya fue eliminado, buscarlo en historial
    if not r.get('nombre_migrante'):
        hist = _exec(db,
            'SELECT folio, nombre_migrante FROM historial_eliminaciones'
            ' WHERE id=(SELECT MIN(id) FROM historial_eliminaciones WHERE solicitado_por=%s)',
            (r.get('solicitado_por'),)
        ).fetchone()
        if hist:
            r['folio'] = hist['folio']
            r['nombre_migrante'] = hist['nombre_migrante']
    for k in ('fecha_solicitud', 'fecha_resolucion'):
        if r.get(k) and hasattr(r[k], 'isoformat'):
            r[k] = r[k].isoformat()
    # Verificar firma si existe
    if r.get('firma_coord') and r.get('coord_pubkey') and r.get('mensaje_firmado'):
        r['firma_valida'] = verificar_firma(r['coord_pubkey'], r['mensaje_firmado'], r['firma_coord'])
        r['firma_coord_corta'] = r['firma_coord'][:48] + '…'
    else:
        r['firma_valida'] = None
        r['firma_coord_corta'] = None
    r.pop('firma_coord', None)
    r.pop('coord_pubkey', None)
    return jsonify(r)


@app.route('/admin/api/actividad_detalle/arco/<string:tipo>/<int:sid>')
@verificar_certificado
def admin_actividad_detalle_arco(tipo, sid):
    if g.rol != 'admin':
        abort(403)
    import json as _json
    db = obtener_db()
    if tipo == 'rect':
        row = _exec(db,
            'SELECT r.*, m.folio, m.nombre AS migrante_nombre'
            ' FROM solicitudes_arco_rect r'
            ' JOIN migrantes m ON m.id=r.migrante_id'
            ' WHERE r.id=%s', (sid,)
        ).fetchone()
        if not row:
            return jsonify(error='Solicitud no encontrada'), 404
        d = dict(row)
        try:
            d['cambios'] = _json.loads(d.get('cambios_json') or '{}')
        except Exception:
            d['cambios'] = {}
        for k in ('fecha_solicitud', 'resuelto_en'):
            if d.get(k) and hasattr(d[k], 'isoformat'):
                d[k] = d[k].isoformat()
        return jsonify(d)
    elif tipo == 'cancelacion':
        row = _exec(db,
            'SELECT s.*, m.folio, m.nombre AS migrante_nombre'
            ' FROM solicitudes_cancelacion_op s'
            ' JOIN migrantes m ON m.id=s.migrante_id'
            ' WHERE s.id=%s', (sid,)
        ).fetchone()
        if not row:
            return jsonify(error='Solicitud no encontrada'), 404
        d = dict(row)
        for k in ('fecha_solicitud', 'resuelto_en'):
            if d.get(k) and hasattr(d[k], 'isoformat'):
                d[k] = d[k].isoformat()
        return jsonify(d)
    else:
        return jsonify(error='Tipo de solicitud inválido'), 400


@app.route('/api/registro_migrante/mis_solicitudes', methods=['GET'])
@verificar_certificado
def api_registro_mis_solicitudes():
    if g.rol not in ('voluntario', 'op'):
        return jsonify(error=f'Rol incorrecto: {g.rol}'), 403
    try:
        db = obtener_db()
        cur = _exec(db,
            "SELECT id, nombre, pais_origen, estado, creado,"
            "       rechazado_por, motivo_rechazo, rechazado_en"
            " FROM solicitudes_registro_migrante"
            " WHERE enviado_por=%s"
            " ORDER BY creado DESC LIMIT 50",
            (g.usuario,)
        )
        filas = cur.fetchall()
    except Exception as e:
        return jsonify(error=str(e)), 500
    resultado = []
    for f in filas:
        r = dict(f)
        for k in ('creado', 'rechazado_en'):
            if r.get(k) and hasattr(r[k], 'isoformat'):
                r[k] = r[k].isoformat()
        resultado.append(r)
    return jsonify(resultado)


@app.route('/api/registro_migrante/<int:sid>/op_validar', methods=['POST'])
@verificar_certificado
def api_registro_op_validar(sid):
    if g.rol != 'op':
        abort(403)
    db = obtener_db()
    cur = _exec(db,
        "SELECT * FROM solicitudes_registro_migrante WHERE id=%s AND estado='pendiente_op'",
        (sid,)
    )
    ticket = cur.fetchone()
    if not ticket:
        return jsonify(error='Solicitud no encontrada o ya procesada'), 404

    accion = request.form.get('accion', '').strip()
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if accion == 'rechazar':
        motivo = request.form.get('motivo', '').strip()
        if not motivo:
            return jsonify(error='El motivo de rechazo es obligatorio'), 400
        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='rechazada', rechazado_por=%s, motivo_rechazo=%s, rechazado_en=%s"
            " WHERE id=%s",
            (g.usuario, motivo, ahora, sid)
        )
        db.commit()
        log_evento('registro_mig_op_rechazado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} rechazado: {motivo}')
        return jsonify(ok=True)

    if accion == 'validar':
        priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
        nombre = ticket['nombre'] or ''
        fnac_str = str(ticket['fecha_nacimiento']) if ticket['fecha_nacimiento'] else ''
        pais_str = ticket['pais_origen'] or ''
        enviado_por = ticket['enviado_por'] or ''
        mensaje_op = (
            f'REGISTRO_OP|{sid}|{nombre}|{fnac_str}|{pais_str}'
            f'|{enviado_por}|{g.usuario}|{ahora}'
        )
        firma_op = firmar_mensaje(priv_pem, mensaje_op)
        if not verificar_firma(pub_pem, mensaje_op, firma_op):
            return jsonify(error='Error al generar firma digital'), 500
        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='pendiente_coord', op_validador=%s, firma_op=%s,"
            "     op_pubkey=%s, mensaje_firmado_op=%s, op_validado_en=%s"
            " WHERE id=%s",
            (g.usuario, firma_op, pub_pem, mensaje_op, ahora, sid)
        )
        db.commit()
        log_evento('registro_mig_op_validado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} validado → pendiente coord. Enviado por: {enviado_por}')
        return jsonify(ok=True)

    return jsonify(error='Acción inválida'), 400


@app.route('/api/registro_migrante/<int:sid>/coord_resolver', methods=['POST'])
@verificar_certificado
def api_registro_coord_resolver(sid):
    if g.rol != 'coord':
        abort(403)
    db = obtener_db()
    import json as _json
    cur = _exec(db,
        "SELECT * FROM solicitudes_registro_migrante WHERE id=%s AND estado='pendiente_coord'",
        (sid,)
    )
    ticket = cur.fetchone()
    if not ticket:
        return jsonify(error='Solicitud no encontrada o ya procesada'), 404

    accion = request.form.get('accion', '').strip()
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if accion == 'rechazar':
        motivo = request.form.get('motivo', '').strip()
        if not motivo:
            return jsonify(error='El motivo de rechazo es obligatorio'), 400
        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='rechazada', rechazado_por=%s, motivo_rechazo=%s, rechazado_en=%s"
            " WHERE id=%s",
            (g.usuario, motivo, ahora, sid)
        )
        db.commit()
        log_evento('registro_mig_coord_rechazado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} rechazado por coord: {motivo}')
        return jsonify(ok=True)

    if accion == 'aprobar':
        priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
        nombre = ticket['nombre'] or ''
        fnac_str = str(ticket['fecha_nacimiento']) if ticket['fecha_nacimiento'] else ''
        pais_str = ticket['pais_origen'] or ''
        op_validador = ticket['op_validador'] or ''
        firma_op_val = ticket['firma_op'] or ''
        firma_op_ancla = firma_op_val[:16] if firma_op_val else ''
        mensaje_coord = (
            f'REGISTRO_COORD|{sid}|{nombre}|{fnac_str}|{pais_str}'
            f'|{op_validador}|{g.usuario}|{ahora}|{firma_op_ancla}'
        )
        firma_coord = firmar_mensaje(priv_pem, mensaje_coord)
        if not verificar_firma(pub_pem, mensaje_coord, firma_coord):
            return jsonify(error='Error al generar firma digital'), 500

        try:
            cur2 = _exec(db,
                'INSERT INTO migrantes'
                ' (folio, nombre, fecha_nacimiento, pais_origen,'
                '  fecha_atencion, genero, departamento_estado, estado_civil,'
                '  grupo_poblacion, telefono_contacto, registrado_por)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (
                    'TEMP', ticket['nombre'], ticket['fecha_nacimiento'],
                    ticket['pais_origen'], ticket['fecha_atencion'],
                    ticket['genero'], ticket['departamento_estado'],
                    ticket['estado_civil'], ticket['grupo_poblacion'],
                    ticket['telefono_contacto'], ticket['enviado_por'],
                )
            )
        except Exception as e:
            return jsonify(error=f'Error al insertar migrante: {e}'), 500

        nuevo_id = cur2.lastrowid
        folio = f'MIG-{datetime.now().strftime("%Y%m")}-{nuevo_id:04d}'
        _exec(db, 'UPDATE migrantes SET folio=%s WHERE id=%s', (folio, nuevo_id))

        firmas = {}
        if ticket['firma_op']:
            firmas['op'] = {
                'usuario': ticket['op_validador'],
                'mensaje': ticket['mensaje_firmado_op'],
                'firma': ticket['firma_op'],
                'pubkey': ticket['op_pubkey'],
            }
        firmas['coord'] = {
            'usuario': g.usuario,
            'mensaje': mensaje_coord,
            'firma': firma_coord,
            'pubkey': pub_pem,
        }
        _exec(db, 'UPDATE migrantes SET firmas_aprobacion=%s WHERE id=%s',
              (_json.dumps(firmas), nuevo_id))

        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='aprobada', coord_aprobador=%s, firma_coord=%s,"
            "     coord_pubkey=%s, mensaje_firmado_coord=%s, coord_aprobado_en=%s"
            " WHERE id=%s",
            (g.usuario, firma_coord, pub_pem, mensaje_coord, ahora, sid)
        )
        db.commit()

        log_evento('registro_mig_coord_aprobado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} aprobado. Folio: {folio}')
        log_evento('registro_mig_insertado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Folio: {folio} · Aprobado por coord desde ticket #{sid}')
        return jsonify(ok=True, folio=folio, migrante_id=nuevo_id)

    return jsonify(error='Acción inválida'), 400


# ── ADMIN: OVERRIDE APROBACIÓN (firma supervisora sobre cualquier pendiente) ──

@app.route('/admin/api/registro_migrante/<int:sid>/admin_resolver', methods=['POST'])
@verificar_certificado
def api_registro_admin_resolver(sid):
    if g.rol != 'admin':
        abort(403)
    import json as _json
    db = obtener_db()
    cur = _exec(db,
        "SELECT * FROM solicitudes_registro_migrante"
        " WHERE id=%s AND estado IN ('pendiente_op','pendiente_coord')",
        (sid,)
    )
    ticket = cur.fetchone()
    if not ticket:
        return jsonify(error='Solicitud no encontrada o ya procesada'), 404

    data = request.get_json(silent=True) or {}
    accion = (data.get('accion') or request.form.get('accion', '')).strip()
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if accion == 'rechazar':
        motivo = (data.get('motivo') or request.form.get('motivo', '')).strip()
        if not motivo:
            return jsonify(error='El motivo de rechazo es obligatorio'), 400
        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='rechazada', rechazado_por=%s, motivo_rechazo=%s, rechazado_en=%s"
            " WHERE id=%s",
            (g.usuario, motivo, ahora, sid)
        )
        db.commit()
        log_evento('registro_mig_coord_rechazado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} rechazado por admin (override): {motivo}')
        return jsonify(ok=True)

    if accion == 'aprobar':
        priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
        nombre    = ticket['nombre'] or ''
        fnac_str  = str(ticket['fecha_nacimiento']) if ticket['fecha_nacimiento'] else ''
        pais_str  = ticket['pais_origen'] or ''
        op_val    = ticket['op_validador'] or ''
        firma_op_ancla = (ticket['firma_op'] or '')[:16]
        mensaje_admin = (
            f'REGISTRO_ADMIN|{sid}|{nombre}|{fnac_str}|{pais_str}'
            f'|{op_val}|{g.usuario}|{ahora}|{firma_op_ancla}'
        )
        firma_admin = firmar_mensaje(priv_pem, mensaje_admin)
        if not verificar_firma(pub_pem, mensaje_admin, firma_admin):
            return jsonify(error='Error al generar firma digital'), 500

        try:
            cur2 = _exec(db,
                'INSERT INTO migrantes'
                ' (folio, nombre, fecha_nacimiento, pais_origen,'
                '  fecha_atencion, genero, departamento_estado, estado_civil,'
                '  grupo_poblacion, telefono_contacto, registrado_por)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (
                    'TEMP', ticket['nombre'], ticket['fecha_nacimiento'],
                    ticket['pais_origen'], ticket['fecha_atencion'],
                    ticket['genero'], ticket['departamento_estado'],
                    ticket['estado_civil'], ticket['grupo_poblacion'],
                    ticket['telefono_contacto'], ticket['enviado_por'],
                )
            )
        except Exception as e:
            return jsonify(error=f'Error al insertar migrante: {e}'), 500

        nuevo_id = cur2.lastrowid
        folio = f'MIG-{datetime.now().strftime("%Y%m")}-{nuevo_id:04d}'
        _exec(db, 'UPDATE migrantes SET folio=%s WHERE id=%s', (folio, nuevo_id))

        firmas = {}
        if ticket['firma_op']:
            firmas['op'] = {
                'usuario': ticket['op_validador'],
                'mensaje': ticket['mensaje_firmado_op'],
                'firma':   ticket['firma_op'],
                'pubkey':  ticket['op_pubkey'],
            }
        firmas['admin'] = {
            'usuario': g.usuario,
            'mensaje': mensaje_admin,
            'firma':   firma_admin,
            'pubkey':  pub_pem,
        }
        _exec(db, 'UPDATE migrantes SET firmas_aprobacion=%s WHERE id=%s',
              (_json.dumps(firmas), nuevo_id))

        _exec(db,
            "UPDATE solicitudes_registro_migrante"
            " SET estado='aprobada', admin_aprobador=%s, firma_admin=%s,"
            "     admin_pubkey=%s, mensaje_firmado_admin=%s, admin_aprobado_en=%s"
            " WHERE id=%s",
            (g.usuario, firma_admin, pub_pem, mensaje_admin, ahora, sid)
        )
        db.commit()

        log_evento('registro_mig_admin_aprobado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Ticket #{sid} aprobado por admin (override). Folio: {folio}')
        log_evento('registro_mig_insertado', usuario=g.usuario, rol=g.rol,
                   detalle=f'Folio: {folio} · Aprobado por admin (override) desde ticket #{sid}')
        return jsonify(ok=True, folio=folio, migrante_id=nuevo_id)

    return jsonify(error='Acción inválida'), 400


# ── BULK RESOLVER (coord y admin) ─────────────────────────────

@app.route('/api/registro_migrante/bulk_coord_resolver', methods=['POST'])
@verificar_certificado
def api_bulk_coord_resolver():
    if g.rol not in ('coord', 'admin'):
        abort(403)
    import json as _json
    data = request.get_json(silent=True) or {}
    ids_raw = data.get('ids') or request.form.get('ids_json', '[]')
    if isinstance(ids_raw, str):
        try:
            ids = _json.loads(ids_raw)
        except Exception:
            return jsonify(error='ids_json inválido'), 400
    else:
        ids = ids_raw
    accion = (data.get('accion') or request.form.get('accion', '')).strip()
    motivo = (data.get('motivo') or request.form.get('motivo', '')).strip()
    if accion not in ('aprobar', 'rechazar'):
        return jsonify(error='Acción inválida'), 400
    if accion == 'rechazar' and not motivo:
        return jsonify(error='El motivo es obligatorio para rechazar'), 400
    if not ids:
        return jsonify(error='Sin IDs'), 400

    db = obtener_db()
    priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
    procesados, errores = 0, []

    for sid in ids:
        try:
            estado_requerido = 'pendiente_coord' if g.rol == 'coord' else "pendiente_coord"
            ticket = _exec(db,
                f"SELECT * FROM solicitudes_registro_migrante WHERE id=%s AND estado='{estado_requerido}'",
                (sid,)
            ).fetchone()
            if not ticket:
                errores.append(f'#{sid}: no encontrado o ya procesado')
                continue
            ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if accion == 'rechazar':
                _exec(db,
                    "UPDATE solicitudes_registro_migrante"
                    " SET estado='rechazada', rechazado_por=%s, motivo_rechazo=%s, rechazado_en=%s"
                    " WHERE id=%s",
                    (g.usuario, motivo, ahora, sid)
                )
                log_evento('registro_mig_coord_rechazado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Ticket #{sid} rechazado: {motivo}')
                procesados += 1
            else:
                nombre = ticket['nombre'] or ''
                fnac_str = str(ticket['fecha_nacimiento']) if ticket['fecha_nacimiento'] else ''
                pais_str = ticket['pais_origen'] or ''
                op_validador = ticket['op_validador'] or ''
                firma_op_val = ticket['firma_op'] or ''
                firma_op_ancla = firma_op_val[:16] if firma_op_val else ''
                mensaje_coord = (
                    f'REGISTRO_COORD|{sid}|{nombre}|{fnac_str}|{pais_str}'
                    f'|{op_validador}|{g.usuario}|{ahora}|{firma_op_ancla}'
                )
                firma_coord = firmar_mensaje(priv_pem, mensaje_coord)
                cur2 = _exec(db,
                    'INSERT INTO migrantes'
                    ' (folio, nombre, fecha_nacimiento, pais_origen,'
                    '  fecha_atencion, genero, departamento_estado, estado_civil,'
                    '  grupo_poblacion, telefono_contacto, registrado_por)'
                    ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    ('TEMP', ticket['nombre'], ticket['fecha_nacimiento'],
                     ticket['pais_origen'], ticket['fecha_atencion'],
                     ticket['genero'], ticket['departamento_estado'],
                     ticket['estado_civil'], ticket['grupo_poblacion'],
                     ticket['telefono_contacto'], ticket['enviado_por'])
                )
                nuevo_id = cur2.lastrowid
                folio = f'MIG-{datetime.now().strftime("%Y%m")}-{nuevo_id:04d}'
                _exec(db, 'UPDATE migrantes SET folio=%s WHERE id=%s', (folio, nuevo_id))
                firmas = {}
                if ticket['firma_op']:
                    firmas['op'] = {'usuario': ticket['op_validador'],
                                    'mensaje': ticket['mensaje_firmado_op'],
                                    'firma': ticket['firma_op'],
                                    'pubkey': ticket['op_pubkey']}
                firmas['coord'] = {'usuario': g.usuario, 'mensaje': mensaje_coord,
                                   'firma': firma_coord, 'pubkey': pub_pem}
                _exec(db, 'UPDATE migrantes SET firmas_aprobacion=%s WHERE id=%s',
                      (_json.dumps(firmas), nuevo_id))
                _exec(db,
                    "UPDATE solicitudes_registro_migrante"
                    " SET estado='aprobada', coord_aprobador=%s, firma_coord=%s,"
                    "     coord_pubkey=%s, mensaje_firmado_coord=%s, coord_aprobado_en=%s"
                    " WHERE id=%s",
                    (g.usuario, firma_coord, pub_pem, mensaje_coord, ahora, sid)
                )
                log_evento('registro_mig_coord_aprobado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Ticket #{sid} aprobado. Folio: {folio}')
                log_evento('registro_mig_insertado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Folio: {folio} · Aprobado por coord desde ticket #{sid}')
                procesados += 1
        except Exception as ex:
            errores.append(f'#{sid}: {ex}')
    db.commit()
    return jsonify(ok=True, procesados=procesados, errores=errores)


@app.route('/api/registro_migrante/bulk_admin_resolver', methods=['POST'])
@verificar_certificado
def api_bulk_admin_resolver():
    if g.rol != 'admin':
        abort(403)
    import json as _json
    data = request.get_json(silent=True) or {}
    ids_raw = data.get('ids') or request.form.get('ids_json', '[]')
    if isinstance(ids_raw, str):
        try:
            ids = _json.loads(ids_raw)
        except Exception:
            return jsonify(error='ids_json inválido'), 400
    else:
        ids = ids_raw
    accion = (data.get('accion') or request.form.get('accion', '')).strip()
    motivo = (data.get('motivo') or request.form.get('motivo', '')).strip()
    if accion not in ('aprobar', 'rechazar'):
        return jsonify(error='Acción inválida'), 400
    if accion == 'rechazar' and not motivo:
        return jsonify(error='El motivo es obligatorio para rechazar'), 400
    if not ids:
        return jsonify(error='Sin IDs'), 400

    db = obtener_db()
    priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
    procesados, errores = 0, []

    for sid in ids:
        try:
            ticket = _exec(db,
                "SELECT * FROM solicitudes_registro_migrante"
                " WHERE id=%s AND estado IN ('pendiente_op','pendiente_coord')",
                (sid,)
            ).fetchone()
            if not ticket:
                errores.append(f'#{sid}: no encontrado o ya procesado')
                continue
            ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if accion == 'rechazar':
                _exec(db,
                    "UPDATE solicitudes_registro_migrante"
                    " SET estado='rechazada', rechazado_por=%s, motivo_rechazo=%s, rechazado_en=%s"
                    " WHERE id=%s",
                    (g.usuario, motivo, ahora, sid)
                )
                log_evento('registro_mig_coord_rechazado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Ticket #{sid} rechazado: {motivo}')
                procesados += 1
            else:
                nombre = ticket['nombre'] or ''
                fnac_str = str(ticket['fecha_nacimiento']) if ticket['fecha_nacimiento'] else ''
                pais_str = ticket['pais_origen'] or ''
                op_validador = ticket['op_validador'] or ''
                firma_op_val = ticket['firma_op'] or ''
                firma_op_ancla = firma_op_val[:16] if firma_op_val else ''
                mensaje_admin = (
                    f'REGISTRO_ADMIN|{sid}|{nombre}|{fnac_str}|{pais_str}'
                    f'|{op_validador}|{g.usuario}|{ahora}|{firma_op_ancla}'
                )
                firma_admin = firmar_mensaje(priv_pem, mensaje_admin)
                cur2 = _exec(db,
                    'INSERT INTO migrantes'
                    ' (folio, nombre, fecha_nacimiento, pais_origen,'
                    '  fecha_atencion, genero, departamento_estado, estado_civil,'
                    '  grupo_poblacion, telefono_contacto, registrado_por)'
                    ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                    ('TEMP', ticket['nombre'], ticket['fecha_nacimiento'],
                     ticket['pais_origen'], ticket['fecha_atencion'],
                     ticket['genero'], ticket['departamento_estado'],
                     ticket['estado_civil'], ticket['grupo_poblacion'],
                     ticket['telefono_contacto'], ticket['enviado_por'])
                )
                nuevo_id = cur2.lastrowid
                folio = f'MIG-{datetime.now().strftime("%Y%m")}-{nuevo_id:04d}'
                _exec(db, 'UPDATE migrantes SET folio=%s WHERE id=%s', (folio, nuevo_id))
                firmas = {}
                if ticket['firma_op']:
                    firmas['op'] = {'usuario': ticket['op_validador'],
                                    'mensaje': ticket['mensaje_firmado_op'],
                                    'firma': ticket['firma_op'],
                                    'pubkey': ticket['op_pubkey']}
                firmas['admin'] = {'usuario': g.usuario, 'mensaje': mensaje_admin,
                                   'firma': firma_admin, 'pubkey': pub_pem}
                _exec(db, 'UPDATE migrantes SET firmas_aprobacion=%s WHERE id=%s',
                      (_json.dumps(firmas), nuevo_id))
                _exec(db,
                    "UPDATE solicitudes_registro_migrante"
                    " SET estado='aprobada', admin_aprobador=%s, firma_admin=%s,"
                    "     admin_pubkey=%s, mensaje_firmado_admin=%s, admin_aprobado_en=%s"
                    " WHERE id=%s",
                    (g.usuario, firma_admin, pub_pem, mensaje_admin, ahora, sid)
                )
                log_evento('registro_mig_admin_aprobado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Ticket #{sid} aprobado (override). Folio: {folio}')
                log_evento('registro_mig_insertado', usuario=g.usuario, rol=g.rol,
                           detalle=f'[Bulk] Folio: {folio} · Aprobado por admin desde ticket #{sid}')
                procesados += 1
        except Exception as ex:
            errores.append(f'#{sid}: {ex}')
    db.commit()
    return jsonify(ok=True, procesados=procesados, errores=errores)


@app.route('/admin/api/actividades_colaboradores')
@verificar_certificado
def api_actividades_colaboradores():
    if g.rol != 'admin':
        abort(403)
    import json as _json
    db = obtener_db()

    _pendiente_de = {
        'pendiente_op':    'Operativo',
        'pendiente_coord': 'Coordinador',
        'aprobada':        'Aprobado',
        'rechazada':       'Rechazado',
    }

    def _fstr(v):
        return str(v)[:19] if v else None

    # Solicitudes de registro de migrante
    rows_reg = _exec(db,
        'SELECT id, nombre, pais_origen, estado, enviado_por,'
        '       op_validador, coord_aprobador,'
        '       rechazado_por, motivo_rechazo, creado'
        ' FROM solicitudes_registro_migrante'
        ' ORDER BY creado DESC LIMIT 200'
    ).fetchall()
    registro = []
    for r in rows_reg:
        d = dict(r)
        d['creado'] = _fstr(d.get('creado'))
        d['pendiente_de'] = _pendiente_de.get(d.get('estado'), d.get('estado', ''))
        registro.append(d)

    # Solicitudes de eliminación
    rows_elim = _exec(db,
        'SELECT s.id, m.folio, m.nombre AS nombre_migrante,'
        '       s.solicitado_por, s.estado, s.fecha_solicitud, s.motivo'
        ' FROM solicitudes_eliminacion s'
        ' JOIN migrantes m ON m.id = s.migrante_id'
        ' ORDER BY s.fecha_solicitud DESC LIMIT 100'
    ).fetchall()
    eliminacion = []
    for r in rows_elim:
        d = dict(r)
        if d.get('fecha_solicitud') and hasattr(d['fecha_solicitud'], 'isoformat'):
            d['fecha_solicitud'] = d['fecha_solicitud'].isoformat()
        eliminacion.append(d)

    # Solicitudes ARCO (rectificación y cancelaciones op)
    rows_rect = _exec(db,
        'SELECT r.id, m.folio, m.nombre AS nombre_migrante,'
        '       r.solicitado_por, r.estado, r.fecha_solicitud AS creado, \'rect\' AS tipo_arco'
        ' FROM solicitudes_arco_rect r'
        ' JOIN migrantes m ON m.id = r.migrante_id'
        ' ORDER BY r.fecha_solicitud DESC LIMIT 50'
    ).fetchall()
    rows_canc = _exec(db,
        'SELECT c.id, m.folio, m.nombre AS nombre_migrante,'
        '       c.solicitado_por, c.estado, c.fecha_solicitud AS creado, \'cancelacion\' AS tipo_arco'
        ' FROM solicitudes_cancelacion_op c'
        ' JOIN migrantes m ON m.id = c.migrante_id'
        ' ORDER BY c.fecha_solicitud DESC LIMIT 50'
    ).fetchall()
    arco = []
    for r in list(rows_rect) + list(rows_canc):
        d = dict(r)
        if d.get('creado') and hasattr(d['creado'], 'isoformat'):
            d['creado'] = d['creado'].isoformat()[:19]
        arco.append(d)
    arco.sort(key=lambda x: x.get('creado') or '', reverse=True)

    return jsonify(registro=registro, eliminacion=eliminacion, arco=arco)


@app.route('/api/migrante/<int:mid>', methods=['GET'])
@verificar_certificado
def api_migrante_detalle(mid):
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    import json as _json
    mig = _exec(db,
        'SELECT id, folio, nombre, fecha_nacimiento, pais_origen,'
        ' fecha_atencion, genero, departamento_estado, estado_civil,'
        ' grupo_poblacion, telefono_contacto, registrado_por,'
        ' firmas_aprobacion, creado'
        ' FROM migrantes WHERE id=%s', (mid,)
    ).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404
    def _fstr(v):
        return v.isoformat()[:10] if hasattr(v, 'isoformat') else (str(v) if v else None)
    m = dict(mig)
    for k in ('fecha_nacimiento', 'fecha_atencion'):
        m[k] = _fstr(m.get(k))
    if m.get('creado'):
        m['creado'] = str(m['creado'])[:19]
    firmas = {}
    if m.get('firmas_aprobacion'):
        try:
            firmas = _json.loads(m['firmas_aprobacion'])
        except Exception:
            pass
    m['firmante_op']    = (firmas.get('op') or {}).get('usuario', '')
    m['firmante_coord'] = (firmas.get('coord') or firmas.get('admin') or {}).get('usuario', '')
    del m['firmas_aprobacion']
    return jsonify(m)


@app.route('/api/migrantes/<int:mid>', methods=['PUT'])
@verificar_certificado
def api_migrantes_editar(mid):
    if g.rol not in ('admin', 'coord'):
        abort(403)

    nombre_pila      = request.form.get('nombre_pila', '').strip()
    primer_apellido  = request.form.get('primer_apellido', '').strip()
    segundo_apellido = request.form.get('segundo_apellido', '').strip()
    fecha_atencion   = request.form.get('fecha_atencion', '').strip() or None
    telefono_contacto = request.form.get('telefono_contacto', '').strip() or None
    genero           = request.form.get('genero', '').strip() or None
    pais_origen      = request.form.get('pais_origen', '').strip()
    departamento_estado = request.form.get('departamento_estado', '').strip() or None
    estado_civil     = request.form.get('estado_civil', '').strip() or None
    fecha_nac        = request.form.get('fecha_nacimiento', '').strip() or None
    grupo_poblacion  = request.form.get('grupo_poblacion', '').strip() or None

    faltantes = [f for f, v in [
        ('nombre_pila', nombre_pila),
        ('primer_apellido', primer_apellido),
        ('pais_origen', pais_origen),
        ('fecha_nacimiento', fecha_nac),
        ('genero', genero),
        ('grupo_poblacion', grupo_poblacion),
    ] if not v]
    if faltantes:
        return jsonify(error=f'Campos obligatorios faltantes: {", ".join(faltantes)}'), 400

    err_fnac = _validar_fecha_rango(fecha_nac, 'Fecha de nacimiento')
    if err_fnac:
        return jsonify(error=err_fnac), 400
    err_fat = _validar_fecha_rango(fecha_atencion, 'Fecha de atención')
    if err_fat:
        return jsonify(error=err_fat), 400

    partes = [nombre_pila, primer_apellido]
    if segundo_apellido:
        partes.append(segundo_apellido)
    nombre = ' '.join(partes)

    db = obtener_db()
    mig = _exec(db,
        'SELECT folio, nombre, fecha_nacimiento, pais_origen,'
        ' fecha_atencion, genero, departamento_estado, estado_civil, grupo_poblacion,'
        ' telefono_contacto FROM migrantes WHERE id=%s', (mid,)
    ).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404

    try:
        _exec(db,
            'UPDATE migrantes SET'
            '  nombre=%s, fecha_nacimiento=%s, pais_origen=%s,'
            '  fecha_atencion=%s, genero=%s, departamento_estado=%s,'
            '  estado_civil=%s, grupo_poblacion=%s, telefono_contacto=%s'
            ' WHERE id=%s',
            (
                nombre, fecha_nac, pais_origen,
                fecha_atencion, genero, departamento_estado,
                estado_civil, grupo_poblacion, telefono_contacto, mid,
            )
        )
    except Exception as e:
        return jsonify(error=f'Error al actualizar en base de datos: {e}'), 500
    db.commit()

    _s = lambda v: v.isoformat()[:10] if hasattr(v, 'isoformat') else (str(v) if v else '')
    comparar = [
        ('nombre',           _s(mig['nombre']),            nombre),
        ('fecha_nacimiento', _s(mig['fecha_nacimiento']),  fecha_nac or ''),
        ('pais_origen',      _s(mig['pais_origen']),       pais_origen),
        ('genero',           _s(mig.get('genero') or ''),  genero or ''),
        ('grupo_poblacion',  _s(mig.get('grupo_poblacion') or ''), grupo_poblacion or ''),
    ]
    cambios = [f'{k}: "{vo}" → "{vn}"' for k, vo, vn in comparar if vo != vn]
    detalle = (f'Folio: {mig["folio"]} · '
               + (' | '.join(cambios) if cambios else 'sin cambios'))
    log_evento('migrante_actualizado', usuario=g.usuario, rol=g.rol, detalle=detalle)
    return jsonify(ok=True)


@app.route('/api/migrantes/<int:mid>', methods=['DELETE'])
@verificar_certificado
def api_migrantes_eliminar(mid):
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    mig = _exec(db, 'SELECT nombre, folio FROM migrantes WHERE id=%s', (mid,)).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404
    _exec(db,
        'INSERT INTO historial_eliminaciones (folio, nombre_migrante, aprobado_por, tipo)'
        ' VALUES (%s,%s,%s,%s)',
        (mig['folio'], mig['nombre'], g.usuario, 'directa')
    )
    _exec(db, 'DELETE FROM migrantes WHERE id=%s', (mid,))
    db.commit()
    log_evento('migrante_eliminado', usuario=g.usuario, rol=g.rol,
               detalle=f'Eliminación directa: {mig["nombre"]} · Folio: {mig["folio"]}')
    return jsonify(ok=True)


@app.route('/api/migrantes/<int:mid>/solicitar_eliminacion', methods=['POST'])
@verificar_certificado
def api_solicitar_eliminacion(mid):
    if g.rol != 'coord':
        abort(403)

    db = obtener_db()
    mig = _exec(db, 'SELECT id, nombre, folio FROM migrantes WHERE id=%s', (mid,)).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404

    pendiente = _exec(db,
        "SELECT id FROM solicitudes_eliminacion"
        " WHERE migrante_id=%s AND estado='pendiente'", (mid,)
    ).fetchone()
    if pendiente:
        return jsonify(error='Ya existe una solicitud pendiente para este registro'), 400

    motivo = request.form.get('motivo', '').strip() or None
    ahora  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Firma EC del coordinador
    coord_keys = _exec(db,
        'SELECT ec_private_key, ec_public_key FROM usuarios WHERE usuario=%s',
        (g.usuario,)
    ).fetchone()
    firma_coord = pub_pem = mensaje_firmado = None
    if coord_keys and coord_keys['ec_private_key']:
        mensaje_firmado = f"SOLICITUD_ELIMINACION|{mid}|{mig['folio']}|{g.usuario}|{ahora}|{motivo or ''}"
        firma_coord     = firmar_mensaje(coord_keys['ec_private_key'], mensaje_firmado)
        pub_pem         = coord_keys['ec_public_key']

    _exec(db,
        'INSERT INTO solicitudes_eliminacion'
        ' (migrante_id, solicitado_por, motivo, firma_coord, coord_pubkey, mensaje_firmado)'
        ' VALUES (%s,%s,%s,%s,%s,%s)',
        (mid, g.usuario, motivo, firma_coord, pub_pem, mensaje_firmado)
    )
    db.commit()

    log_evento('migrante_solicitud_eliminacion', usuario=g.usuario, rol=g.rol,
               detalle=f'Solicitud eliminación: {mig["nombre"]} · Folio: {mig["folio"]} · Motivo: {motivo or "sin motivo"} · Firmada: {"sí" if firma_coord else "no"}')
    return jsonify(ok=True)


@app.route('/api/mis_solicitudes_eliminacion')
@verificar_certificado
def api_mis_solicitudes_eliminacion():
    if g.rol != 'coord':
        abort(403)
    db = obtener_db()
    vivas = _exec(db,
        'SELECT s.id, m.folio, m.nombre AS nombre_migrante,'
        '  s.motivo, s.estado, s.fecha_solicitud, s.resuelto_por, s.es_arco'
        ' FROM solicitudes_eliminacion s'
        ' JOIN migrantes m ON m.id = s.migrante_id'
        " WHERE s.solicitado_por=%s AND s.estado IN ('pendiente','rechazada')"
        ' ORDER BY s.fecha_solicitud DESC',
        (g.usuario,)
    ).fetchall()
    aprobadas = _exec(db,
        'SELECT id, folio, nombre_migrante, motivo, aprobado_por, fecha_eliminacion, es_arco'
        ' FROM historial_eliminaciones'
        " WHERE solicitado_por=%s AND tipo='solicitada'"
        ' ORDER BY fecha_eliminacion DESC',
        (g.usuario,)
    ).fetchall()
    resultado = []
    for r in vivas:
        d = dict(r)
        if d.get('fecha_solicitud') and hasattr(d['fecha_solicitud'], 'isoformat'):
            d['fecha_solicitud'] = d['fecha_solicitud'].isoformat()
        d['tipo_origen'] = 'viva'
        resultado.append(d)
    for r in aprobadas:
        d = dict(r)
        if d.get('fecha_eliminacion') and hasattr(d['fecha_eliminacion'], 'isoformat'):
            d['fecha_eliminacion'] = d['fecha_eliminacion'].isoformat()
        d['estado'] = 'aprobada'
        d['tipo_origen'] = 'historial'
        resultado.append(d)
    resultado.sort(
        key=lambda x: x.get('fecha_solicitud') or x.get('fecha_eliminacion') or '',
        reverse=True
    )
    return jsonify(resultado)


@app.route('/admin/api/solicitudes')
@verificar_certificado
def api_solicitudes_lista():
    if g.rol not in ('admin', 'coord'):
        abort(403)
    db = obtener_db()
    rows = _exec(db,
        'SELECT s.id, s.migrante_id, m.folio, m.nombre AS migrante_nombre,'
        '  s.solicitado_por, s.motivo, s.estado, s.fecha_solicitud,'
        '  s.solicitante_op, s.firma_coord, s.es_arco'
        ' FROM solicitudes_eliminacion s'
        ' JOIN migrantes m ON m.id = s.migrante_id'
        " WHERE s.estado='pendiente'"
        ' ORDER BY s.fecha_solicitud ASC'
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('fecha_solicitud') and hasattr(d['fecha_solicitud'], 'isoformat'):
            d['fecha_solicitud'] = d['fecha_solicitud'].isoformat()
        result.append(d)
    return jsonify(result)


@app.route('/admin/api/solicitudes/<int:sid>/resolver', methods=['POST'])
@verificar_certificado
def api_resolver_solicitud(sid):
    if g.rol != 'admin':
        abort(403)

    aprobar = request.form.get('aprobar', '0') == '1'
    db = obtener_db()
    sol = _exec(db, 'SELECT * FROM solicitudes_eliminacion WHERE id=%s', (sid,)).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    if sol['estado'] != 'pendiente':
        return jsonify(error='La solicitud ya fue resuelta'), 400

    nuevo_estado = 'aprobada' if aprobar else 'rechazada'
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _exec(db,
        'UPDATE solicitudes_eliminacion'
        ' SET estado=%s, resuelto_por=%s, fecha_resolucion=%s WHERE id=%s',
        (nuevo_estado, g.usuario, ahora, sid)
    )
    if aprobar:
        mig = _exec(db, 'SELECT nombre, folio FROM migrantes WHERE id=%s',
                    (sol['migrante_id'],)).fetchone()
        if mig:
            # Archivar en historial antes de que el CASCADE elimine la solicitud
            _exec(db,
                'INSERT INTO historial_eliminaciones'
                ' (folio, nombre_migrante, fecha_solicitud, solicitado_por, motivo,'
                '  firma_coord, coord_pubkey, mensaje_firmado, aprobado_por, tipo, es_arco)'
                ' VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (mig['folio'], mig['nombre'],
                 sol['fecha_solicitud'], sol['solicitado_por'], sol['motivo'],
                 sol['firma_coord'], sol['coord_pubkey'], sol['mensaje_firmado'],
                 g.usuario, 'solicitada', sol.get('es_arco', 0) or 0)
            )
        _exec(db, 'DELETE FROM migrantes WHERE id=%s', (sol['migrante_id'],))
        db.commit()
        if mig:
            log_evento('migrante_eliminado', usuario=g.usuario, rol=g.rol,
                       detalle=f'Eliminación aprobada: {mig["nombre"]} · Folio: {mig["folio"]} · Sol. por: {sol["solicitado_por"]}')
    else:
        db.commit()
    return jsonify(ok=True, accion=nuevo_estado)


# ── DERECHOS ARCO ─────────────────────────────────────────────
# Whitelist de campos editables por rectificación (previene inyección SQL)
CAMPOS_RECT = {
    'nombre', 'fecha_nacimiento', 'pais_origen', 'departamento_estado',
    'estado_civil', 'genero', 'grupo_poblacion', 'telefono_contacto', 'fecha_atencion'
}

@app.route('/api/arco/buscar')
@verificar_certificado
def arco_buscar():
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    db = obtener_db()
    rows = _exec(db,
        'SELECT id, folio, nombre, pais_origen, fecha_atencion, genero, grupo_poblacion'
        ' FROM migrantes WHERE nombre LIKE %s OR folio LIKE %s'
        ' ORDER BY nombre LIMIT 20',
        (f'%{q}%', f'%{q}%')
    ).fetchall()
    def _fstr(v):
        return v.isoformat()[:10] if hasattr(v, 'isoformat') else (str(v) if v else None)
    return jsonify([{**dict(r), 'fecha_atencion': _fstr(r['fecha_atencion'])} for r in rows])


@app.route('/api/arco/acceso/<int:mid>')
@verificar_certificado
def arco_acceso(mid):
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    mig = _exec(db,
        'SELECT id, folio, nombre, fecha_nacimiento, pais_origen,'
        ' fecha_atencion, genero, departamento_estado, estado_civil,'
        ' grupo_poblacion, telefono_contacto, registrado_por, creado'
        ' FROM migrantes WHERE id=%s', (mid,)
    ).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404
    def _fstr(v):
        return v.isoformat()[:10] if hasattr(v, 'isoformat') else (str(v) if v else None)
    m = dict(mig)
    for k in ('fecha_nacimiento', 'fecha_atencion'):
        m[k] = _fstr(m.get(k))
    if m.get('creado'):
        m['creado'] = m['creado'].isoformat()[:10]
    log_evento('arco_acceso', usuario=g.usuario, rol=g.rol,
               detalle=f'Acceso ARCO id={mid} folio={m["folio"]}')
    return jsonify(m)


@app.route('/api/arco/rectificacion', methods=['POST'])
@verificar_certificado
def arco_rectificacion():
    import json as _json
    if g.rol not in ('op', 'coord'):
        abort(403)
    migrante_id = request.form.get('migrante_id', '').strip()
    cambios_raw = request.form.get('cambios_json', '{}')
    if not migrante_id:
        return jsonify(error='Migrante requerido'), 400
    try:
        cambios = _json.loads(cambios_raw)
    except Exception:
        return jsonify(error='Datos inválidos'), 400
    cambios = {k: v for k, v in cambios.items() if k in CAMPOS_RECT}
    if not cambios:
        return jsonify(error='No hay cambios válidos'), 400
    db = obtener_db()
    mig = _exec(db, 'SELECT folio, nombre FROM migrantes WHERE id=%s', (migrante_id,)).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404

    # Coordinador: aplica los cambios de inmediato
    if g.rol == 'coord':
        sets = ', '.join(f'{k}=%s' for k in cambios)
        vals = list(cambios.values()) + [migrante_id]
        _exec(db, f'UPDATE migrantes SET {sets} WHERE id=%s', vals)
        db.commit()
        log_evento('arco_rect_aprobada', usuario=g.usuario, rol=g.rol,
                   detalle=f'Rectificación inmediata ARCO folio={mig["folio"]}')
        return jsonify(ok=True, aplicado=True)

    # Operativo: crea solicitud pendiente para el coordinador
    pendiente = _exec(db,
        "SELECT id FROM solicitudes_arco_rect WHERE migrante_id=%s AND estado='pendiente'",
        (migrante_id,)
    ).fetchone()
    if pendiente:
        return jsonify(error='Ya existe una solicitud pendiente para este registro'), 400
    _exec(db,
        'INSERT INTO solicitudes_arco_rect (migrante_id, solicitado_por, cambios_json)'
        ' VALUES (%s,%s,%s)',
        (migrante_id, g.usuario, _json.dumps(cambios))
    )
    db.commit()
    log_evento('arco_rect_solicitada', usuario=g.usuario, rol=g.rol,
               detalle=f'Rectificación ARCO folio={mig["folio"]}')
    return jsonify(ok=True, aplicado=False)


@app.route('/api/arco/solicitudes_rect')
@verificar_certificado
def arco_solicitudes_rect():
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    if g.rol == 'op':
        rows = _exec(db,
            'SELECT r.id, r.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  r.solicitado_por, r.cambios_json, r.estado, r.motivo_rechazo,'
            '  r.fecha_solicitud'
            ' FROM solicitudes_arco_rect r'
            ' JOIN migrantes m ON m.id=r.migrante_id'
            ' WHERE r.solicitado_por=%s ORDER BY r.fecha_solicitud DESC LIMIT 50',
            (g.usuario,)
        ).fetchall()
    else:
        rows = _exec(db,
            'SELECT r.id, r.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  r.solicitado_por, r.cambios_json, r.estado, r.motivo_rechazo,'
            '  r.fecha_solicitud'
            ' FROM solicitudes_arco_rect r'
            ' JOIN migrantes m ON m.id=r.migrante_id'
            " WHERE r.estado='pendiente' ORDER BY r.fecha_solicitud ASC"
        ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('fecha_solicitud') and hasattr(d['fecha_solicitud'], 'isoformat'):
            d['fecha_solicitud'] = d['fecha_solicitud'].isoformat()
        result.append(d)
    return jsonify(result)


@app.route('/api/arco/solicitudes_rect/<int:sid>/resolver', methods=['POST'])
@verificar_certificado
def arco_resolver_rect(sid):
    import json as _json
    if g.rol not in ('coord', 'admin'):
        abort(403)
    aprobar = request.form.get('accion') == 'aprobar'
    motivo  = request.form.get('motivo', '').strip() or None
    db = obtener_db()
    sol = _exec(db,
        'SELECT r.*, m.folio FROM solicitudes_arco_rect r'
        ' JOIN migrantes m ON m.id=r.migrante_id WHERE r.id=%s', (sid,)
    ).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    if sol['estado'] != 'pendiente':
        return jsonify(error='La solicitud ya fue resuelta'), 400
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if aprobar:
        try:
            cambios = _json.loads(sol['cambios_json'] or '{}')
        except Exception:
            cambios = {}
        cambios = {k: v for k, v in cambios.items() if k in CAMPOS_RECT}
        if cambios:
            sets = ', '.join(f'{k}=%s' for k in cambios)
            vals = list(cambios.values()) + [sol['migrante_id']]
            _exec(db, f'UPDATE migrantes SET {sets} WHERE id=%s', vals)
    _exec(db,
        'UPDATE solicitudes_arco_rect'
        ' SET estado=%s, resuelto_por=%s, resuelto_en=%s, motivo_rechazo=%s'
        ' WHERE id=%s',
        ('aprobada' if aprobar else 'rechazada', g.usuario, ahora, motivo, sid)
    )
    db.commit()
    log_evento('arco_rect_aprobada' if aprobar else 'arco_rect_rechazada',
               usuario=g.usuario, rol=g.rol,
               detalle=f'Folio {sol["folio"]}: {"aprobada" if aprobar else "rechazada"}')
    return jsonify(ok=True, accion='aprobada' if aprobar else 'rechazada')


# Operativo → Coordinador: solicitar cancelación
@app.route('/api/arco/cancelacion_op', methods=['POST'])
@verificar_certificado
def arco_cancelacion_op():
    if g.rol != 'op':
        abort(403)
    mid_raw = request.form.get('migrante_id', '').strip()
    motivo  = request.form.get('motivo', '').strip() or None
    if not mid_raw or not mid_raw.isdigit():
        return jsonify(error='ID de migrante inválido'), 400
    mid = int(mid_raw)
    db = obtener_db()
    try:
        mig = _exec(db, 'SELECT nombre, folio FROM migrantes WHERE id=%s', (mid,)).fetchone()
        if not mig:
            return jsonify(error='Migrante no encontrado'), 404
        pendiente = _exec(db,
            "SELECT id FROM solicitudes_cancelacion_op WHERE migrante_id=%s AND estado='pendiente'",
            (mid,)
        ).fetchone()
        if pendiente:
            return jsonify(error='Ya existe una solicitud pendiente para este registro'), 400
        _exec(db,
            'INSERT INTO solicitudes_cancelacion_op (migrante_id, solicitado_por, motivo)'
            ' VALUES (%s,%s,%s)',
            (mid, g.usuario, motivo)
        )
        db.commit()
    except Exception as e:
        db.rollback()
        return jsonify(error=f'Error al guardar solicitud: {e}'), 500
    log_evento('arco_cancel_op_solicitada', usuario=g.usuario, rol=g.rol,
               detalle=f'Op solicita cancelación: {mig["nombre"]} · Folio: {mig["folio"]}')
    return jsonify(ok=True)


# Coordinador: ver solicitudes de cancelación enviadas por operativos
@app.route('/api/arco/solicitudes_cancelacion_op')
@verificar_certificado
def arco_solicitudes_cancelacion_op():
    if g.rol != 'coord':
        abort(403)
    db = obtener_db()
    rows = _exec(db,
        'SELECT s.id, s.migrante_id, m.folio, m.nombre AS migrante_nombre,'
        '  s.solicitado_por, s.motivo, s.estado, s.fecha_solicitud'
        ' FROM solicitudes_cancelacion_op s'
        ' JOIN migrantes m ON m.id = s.migrante_id'
        " WHERE s.estado='pendiente'"
        ' ORDER BY s.fecha_solicitud ASC'
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if d.get('fecha_solicitud') and hasattr(d['fecha_solicitud'], 'isoformat'):
            d['fecha_solicitud'] = d['fecha_solicitud'].isoformat()
        result.append(d)
    return jsonify(result)


# Op: ver mis solicitudes de cancelación (historial)
@app.route('/api/arco/mis_solicitudes_cancelacion')
@verificar_certificado
def arco_mis_solicitudes_cancelacion():
    if g.rol != 'op':
        abort(403)
    db = obtener_db()
    rows = _exec(db,
        'SELECT s.id, s.migrante_id, m.folio, m.nombre AS migrante_nombre,'
        '  s.motivo, s.estado, s.fecha_solicitud, s.resuelto_por, s.resuelto_en'
        ' FROM solicitudes_cancelacion_op s'
        ' JOIN migrantes m ON m.id = s.migrante_id'
        ' WHERE s.solicitado_por=%s ORDER BY s.fecha_solicitud DESC LIMIT 50',
        (g.usuario,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for f in ('fecha_solicitud', 'resuelto_en'):
            if d.get(f) and hasattr(d[f], 'isoformat'):
                d[f] = d[f].isoformat()
        result.append(d)
    return jsonify(result)


# Coordinador → Admin: aprobar (firma EC + escala) o rechazar cancelación de op
@app.route('/api/arco/solicitudes_cancelacion_op/<int:sid>/resolver', methods=['POST'])
@verificar_certificado
def arco_resolver_cancelacion_op(sid):
    if g.rol != 'coord':
        abort(403)
    aprobar = request.form.get('accion') == 'aprobar'
    db = obtener_db()
    sol = _exec(db,
        'SELECT s.*, m.folio, m.nombre AS migrante_nombre'
        ' FROM solicitudes_cancelacion_op s'
        ' JOIN migrantes m ON m.id=s.migrante_id WHERE s.id=%s', (sid,)
    ).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    if sol['estado'] != 'pendiente':
        return jsonify(error='La solicitud ya fue resuelta'), 400

    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    nuevo_estado = 'aprobada' if aprobar else 'rechazada'
    _exec(db,
        'UPDATE solicitudes_cancelacion_op'
        ' SET estado=%s, resuelto_por=%s, resuelto_en=%s WHERE id=%s',
        (nuevo_estado, g.usuario, ahora, sid)
    )

    if aprobar:
        # Genera firma EC del coordinador
        priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
        mensaje = (f"CANCELACION|{sol['migrante_id']}|{sol['folio']}"
                   f"|{sol['solicitado_por']}|{g.usuario}|{ahora}|{sol['motivo'] or ''}")
        firma = firmar_mensaje(priv_pem, mensaje)
        # Verifica que la firma sea válida antes de guardar
        if not verificar_firma(pub_pem, mensaje, firma):
            db.rollback()
            return jsonify(error='Error al generar firma digital'), 500
        # Escala a solicitudes_eliminacion para admin
        _exec(db,
            'INSERT INTO solicitudes_eliminacion'
            ' (migrante_id, solicitado_por, motivo, solicitante_op, firma_coord, coord_pubkey, mensaje_firmado, es_arco)'
            ' VALUES (%s,%s,%s,%s,%s,%s,%s,1)',
            (sol['migrante_id'], g.usuario, sol['motivo'],
             sol['solicitado_por'], firma, pub_pem, mensaje)
        )
        db.commit()
        log_evento('arco_cancel_coord_firmada', usuario=g.usuario, rol=g.rol,
                   detalle=f'Cancelación firmada y escalada al admin: {sol["migrante_nombre"]} · Folio: {sol["folio"]}')
    else:
        db.commit()
        log_evento('arco_cancel_coord_rechazada', usuario=g.usuario, rol=g.rol,
                   detalle=f'Cancelación rechazada por coord: {sol["migrante_nombre"]}')

    return jsonify(ok=True, accion=nuevo_estado)


# Admin: verificar firma EC de una solicitud de eliminación
@app.route('/api/arco/verificar_firma/<int:sol_id>')
@verificar_certificado
def arco_verificar_firma(sol_id):
    if g.rol != 'admin':
        abort(403)
    db = obtener_db()
    sol = _exec(db,
        'SELECT firma_coord, coord_pubkey, mensaje_firmado, solicitado_por, solicitante_op'
        ' FROM solicitudes_eliminacion WHERE id=%s', (sol_id,)
    ).fetchone()
    if not sol:
        return jsonify(error='Solicitud no encontrada'), 404
    if not sol['firma_coord']:
        return jsonify(firmada=False, mensaje='Esta solicitud no tiene firma digital')
    valida = verificar_firma(sol['coord_pubkey'], sol['mensaje_firmado'], sol['firma_coord'])
    return jsonify(
        firmada=True,
        valida=valida,
        coord=sol['solicitado_por'],
        solicitante_op=sol['solicitante_op'],
        mensaje=sol['mensaje_firmado'],
        pubkey=sol['coord_pubkey'],
    )


# Coordinador: cancelación directa propia (sin escalado de op) — mantiene flujo existente
@app.route('/api/arco/cancelacion/<int:mid>', methods=['POST'])
@verificar_certificado
def arco_cancelacion(mid):
    if g.rol != 'coord':
        abort(403)
    db = obtener_db()
    mig = _exec(db, 'SELECT nombre, folio FROM migrantes WHERE id=%s', (mid,)).fetchone()
    if not mig:
        return jsonify(error='Migrante no encontrado'), 404
    pendiente = _exec(db,
        "SELECT id FROM solicitudes_eliminacion WHERE migrante_id=%s AND estado='pendiente'",
        (mid,)
    ).fetchone()
    if pendiente:
        return jsonify(error='Ya existe una solicitud de cancelación pendiente'), 400
    motivo = request.form.get('motivo', '').strip() or None
    priv_pem, pub_pem = _obtener_o_crear_claves_ec(db, g.usuario)
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    mensaje = (f"CANCELACION|{mid}|{mig['folio']}"
               f"||{g.usuario}|{ahora}|{motivo or ''}")
    firma = firmar_mensaje(priv_pem, mensaje)
    _exec(db,
        'INSERT INTO solicitudes_eliminacion'
        ' (migrante_id, solicitado_por, motivo, firma_coord, coord_pubkey, mensaje_firmado, es_arco)'
        ' VALUES (%s,%s,%s,%s,%s,%s,1)',
        (mid, g.usuario, motivo, firma, pub_pem, mensaje)
    )
    db.commit()
    log_evento('arco_cancel_coord_firmada', usuario=g.usuario, rol=g.rol,
               detalle=f'Cancelación directa firmada: {mig["nombre"]} · Folio: {mig["folio"]}')
    return jsonify(ok=True)


# Historial unificado de solicitudes ARCO para op/coord/admin
@app.route('/api/arco/historial_completo')
@verificar_certificado
def arco_historial_completo():
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    resultado = []

    def _fstr(v):
        return v.isoformat() if hasattr(v, 'isoformat') else (str(v) if v else None)

    # Solicitudes de rectificación
    if g.rol == 'op':
        rect_rows = _exec(db,
            'SELECT r.id, r.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  r.solicitado_por, r.estado, r.fecha_solicitud, r.resuelto_por, r.resuelto_en,'
            '  r.motivo_rechazo'
            ' FROM solicitudes_arco_rect r'
            ' JOIN migrantes m ON m.id=r.migrante_id'
            ' WHERE r.solicitado_por=%s ORDER BY r.fecha_solicitud DESC LIMIT 100',
            (g.usuario,)
        ).fetchall()
    else:
        rect_rows = _exec(db,
            'SELECT r.id, r.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  r.solicitado_por, r.estado, r.fecha_solicitud, r.resuelto_por, r.resuelto_en,'
            '  r.motivo_rechazo'
            ' FROM solicitudes_arco_rect r'
            ' JOIN migrantes m ON m.id=r.migrante_id'
            ' ORDER BY r.fecha_solicitud DESC LIMIT 200'
        ).fetchall()
    for r in rect_rows:
        d = dict(r)
        d['tipo'] = 'rectificacion'
        d['es_arco'] = 1
        d['firma_coord'] = None
        d['mensaje_firmado'] = None
        for f in ('fecha_solicitud', 'resuelto_en'):
            if d.get(f):
                d[f] = _fstr(d[f])
        resultado.append(d)

    # Solicitudes de cancelación op→coord
    if g.rol == 'op':
        cancel_op_rows = _exec(db,
            'SELECT s.id, s.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  s.solicitado_por, s.estado, s.fecha_solicitud, s.resuelto_por, s.resuelto_en, NULL AS motivo_rechazo'
            ' FROM solicitudes_cancelacion_op s'
            ' JOIN migrantes m ON m.id=s.migrante_id'
            ' WHERE s.solicitado_por=%s ORDER BY s.fecha_solicitud DESC LIMIT 100',
            (g.usuario,)
        ).fetchall()
    else:
        cancel_op_rows = _exec(db,
            'SELECT s.id, s.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  s.solicitado_por, s.estado, s.fecha_solicitud, s.resuelto_por, s.resuelto_en, NULL AS motivo_rechazo'
            ' FROM solicitudes_cancelacion_op s'
            ' JOIN migrantes m ON m.id=s.migrante_id'
            ' ORDER BY s.fecha_solicitud DESC LIMIT 200'
        ).fetchall()
    for r in cancel_op_rows:
        d = dict(r)
        d['tipo'] = 'cancelacion'
        d['es_arco'] = 1
        d['firma_coord'] = None
        d['mensaje_firmado'] = None
        for f in ('fecha_solicitud', 'resuelto_en'):
            if d.get(f):
                d[f] = _fstr(d[f])
        resultado.append(d)

    # Solicitudes de eliminación ARCO (coord y admin ven las firmadas)
    if g.rol in ('coord', 'admin'):
        elim_rows = _exec(db,
            'SELECT e.id, e.migrante_id, m.folio, m.nombre AS migrante_nombre,'
            '  e.solicitado_por, e.estado, e.fecha_solicitud, e.resuelto_por, e.fecha_resolucion AS resuelto_en,'
            '  e.solicitante_op, e.firma_coord, e.mensaje_firmado, NULL AS motivo_rechazo'
            ' FROM solicitudes_eliminacion e'
            ' JOIN migrantes m ON m.id=e.migrante_id'
            ' WHERE e.es_arco=1 ORDER BY e.fecha_solicitud DESC LIMIT 200'
        ).fetchall()
        for r in elim_rows:
            d = dict(r)
            d['tipo'] = 'cancelacion_firmada'
            d['es_arco'] = 1
            for f in ('fecha_solicitud', 'resuelto_en'):
                if d.get(f):
                    d[f] = _fstr(d[f])
            resultado.append(d)

    resultado.sort(key=lambda x: x.get('fecha_solicitud') or '', reverse=True)
    return jsonify(resultado)


# Conteo de pendientes para badge de sidebar
@app.route('/api/pendientes_count')
@verificar_certificado
def api_pendientes_count():
    if g.rol not in ('admin', 'coord', 'op'):
        abort(403)
    db = obtener_db()
    total = 0
    if g.rol == 'op':
        row = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_registro_migrante WHERE estado='pendiente_op'"
        ).fetchone()
        total = row['c'] if row else 0
    elif g.rol == 'coord':
        row = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_registro_migrante WHERE estado='pendiente_coord'"
        ).fetchone()
        total = row['c'] if row else 0
        r_rect = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_arco_rect WHERE estado='pendiente'"
        ).fetchone()
        r_cancel = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_cancelacion_op WHERE estado='pendiente'"
        ).fetchone()
        arco_total = (r_rect['c'] if r_rect else 0) + (r_cancel['c'] if r_cancel else 0)
        return jsonify(total=total, arco_total=arco_total)
    else:
        r1 = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_registro_migrante WHERE estado='pendiente_coord'"
        ).fetchone()
        r2 = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_pwd WHERE estado='pendiente'"
        ).fetchone()
        r3 = _exec(db,
            "SELECT COUNT(*) c FROM solicitudes_eliminacion WHERE estado='pendiente'"
        ).fetchone()
        total = (r1['c'] if r1 else 0) + (r2['c'] if r2 else 0) + (r3['c'] if r3 else 0)
    return jsonify(total=total)


# ── ARRANQUE ──────────────────────────────────────────────────
inicializar_ca()
inicializar_cert_servidor()
inicializar_db()

if __name__ == '__main__':
    import ssl

    print('\n── Casa Monarca Demo ─────────────────────────────')
    if MTLS_ENABLED:
        print('   URL:      https://localhost:5001')
        print('   Modo:     TLS ACTIVO — solo cert de servidor (sin cliente)')
    else:
        print('   URL:      http://localhost:5001')
        print('   Modo:     DEBUG — TLS desactivado')
    print('   Usuario:  admin    Contraseña: admin123')
    print('   Usuario:  coord    Contraseña: coord123')
    print('─────────────────────────────────────────────────\n')

    if MTLS_ENABLED:
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        ctx.load_cert_chain(SERVER_CERT_PATH, SERVER_KEY_PATH)
        ctx.load_verify_locations(CA_CERT_PATH)

        # 🔴 Cambio: ya no exigimos certificado de cliente
        ctx.verify_mode = ssl.CERT_NONE

        app.run(
            debug=False,
            port=5001,
            ssl_context=ctx,
            host='127.0.0.1',
            threaded=True
        )
    else:
        app.run(debug=True, port=5001)
        